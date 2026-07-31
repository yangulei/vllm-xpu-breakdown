# SPDX-License-Identifier: Apache-2.0
"""Excel serialization of Shape Matrix rows (the human/offline exchange format).

Kept apart from :mod:`breakdown.shape_matrix` so the pipeline itself never
depends on openpyxl formatting, and so a matrix produced on one box can be
carried to another (e.g. the CUDA reference machine) and re-read with
the benchmark pipeline.
"""
from __future__ import annotations

import io
from typing import Any, Sequence

from breakdown.shape_matrix import MATRIX_HEADERS


def sheet_name_for(model_id: str) -> str:
    model_short = model_id.split("/")[-1] if "/" in model_id else model_id
    return model_short[:31].replace("[", "").replace("]", "")


def write_workbook(rows: Sequence[dict[str, Any]],
                   info_rows: Sequence[tuple[str, Any]],
                   sheet_name: str) -> bytes:
    """Serialize matrix rows + a provenance sheet into an .xlsx byte string."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E",
                              fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

    ws = wb.active
    ws.title = sheet_name
    for col, hdr in enumerate(MATRIX_HEADERS, 1):
        c = ws.cell(1, col, hdr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    for r in rows:
        for col, hdr in enumerate(MATRIX_HEADERS, 1):
            ws.cell(row, col, r.get(hdr))
        for c in range(1, len(MATRIX_HEADERS) + 1):
            ws.cell(row, c).border = thin_border
        row += 1

    # AutoFit column widths by sampling header + first/last 100 data rows
    sample_rows = list(range(1, min(row, 102)))
    if row > 202:
        sample_rows += list(range(row - 100, row))
    elif row > 102:
        sample_rows += list(range(102, row))
    for col_idx in range(1, len(MATRIX_HEADERS) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for r_idx in sample_rows:
            val = ws.cell(r_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(MATRIX_HEADERS))}{row - 1}"

    info = wb.create_sheet("Info")
    for r_idx, (k, v) in enumerate(info_rows, 1):
        kc = info.cell(r_idx, 1, k)
        kc.font = Font(bold=True, size=10)
        kc.alignment = Alignment(vertical="top")
        vc = info.cell(r_idx, 2, "" if v is None else str(v))
        vc.alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
