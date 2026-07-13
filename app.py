#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""
vLLM Ops/Kernels Breakdown — Web Application.

Interactive web UI for profiling vLLM inference on Intel XPU and visualizing
the op dispatch breakdown.

Usage:
    python app.py [--port 8080] [--host 0.0.0.0]
"""
from __future__ import annotations

import argparse
import functools
import io
import json
import logging
import os
import sys
import threading
import time
import traceback
from dataclasses import asdict
from pathlib import Path
from typing import Any

from flask import Flask, Response, jsonify, request, send_file, send_from_directory

# Force spawn for multiprocessing so vLLM's EngineCore doesn't hit
# "Cannot re-initialize CUDA in forked subprocess".
import multiprocessing
try:
    multiprocessing.set_start_method("spawn", force=True)
except RuntimeError:
    pass  # Already set

os.environ.setdefault("VLLM_WORKER_MULTIPROC_METHOD", "spawn")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from breakdown.analyzer import (
    AnalyzedOp,
    analyze_ops,
    dtype_size,
    estimate_flops,
    estimate_memory,
)
from breakdown.classifier import Backend, classify_op
from breakdown.graph_from_trace import build_graph_from_trace
from breakdown.trace_parser import _detect_device_via_torch
from breakdown.model_info import (
    fetch_model_config,
    get_dim_symbols,
    min_profile_layers,
    summarize_config,
)
from breakdown.registry import ALL_VLLM_XPU_OPS

app = Flask(__name__, static_folder="static")


# Cached at import time — won't change during server lifetime.
_DEVICE = _detect_device_via_torch() or "xpu"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("vllm_xpu_breakdown")

# ---- Config Cache ----
# Persists successfully loaded model configs to disk so they appear as suggestions.

_CONFIG_CACHE_DIR = Path(__file__).parent / "output" / "config_cache"
_CONFIG_CACHE_DIR.mkdir(parents=True, exist_ok=True)
_config_cache_lock = threading.Lock()


def _cache_key(model_id: str) -> str:
    """Convert model_id to a safe filename."""
    return model_id.replace("/", "__")


def _save_config_cache(model_id: str, config: dict[str, Any]) -> None:
    """Persist config.json to disk cache."""
    key = _cache_key(model_id)
    path = _CONFIG_CACHE_DIR / f"{key}.json"
    with _config_cache_lock:
        path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")


def _load_cached_model_ids() -> list[str]:
    """Return list of model IDs that have been successfully cached."""
    ids: list[str] = []
    with _config_cache_lock:
        for p in sorted(_CONFIG_CACHE_DIR.glob("*.json")):
            ids.append(p.stem.replace("__", "/"))
    return ids


def _load_cached_config(model_id: str) -> dict[str, Any] | None:
    """Load a cached config from disk, or None if not cached."""
    key = _cache_key(model_id)
    path = _CONFIG_CACHE_DIR / f"{key}.json"
    if path.exists():
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return None


# Global state for async profiling
_profile_state = {
    "status": "idle",   # idle | running | done | error
    "result": None,
    "error": None,
    "model_id": None,
    "settings": None,
}
_profile_lock = threading.Lock()


# ---- Static file serving ----

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


# ---- Model API ----

@app.route("/api/cached-models")
def get_cached_models():
    """Return list of model IDs whose config.json has been cached locally."""
    return jsonify({"ok": True, "models": _load_cached_model_ids()})


@app.route("/api/model/<path:model_id>")
def get_model_config(model_id: str):
    """Fetch config.json from HuggingFace (or cache) and return summary."""
    try:
        # Try cache first
        config = _load_cached_config(model_id)
        if config is None:
            config = fetch_model_config(model_id)
        # Cache on success
        _save_config_cache(model_id, config)
        summary = summarize_config(config)
        return jsonify({
            "ok": True,
            "config": config,
            "summary": summary,
            "min_profile_layers": min_profile_layers(summary),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ---- Profile API ----


def _set_num_hidden_layers(hf_config, n: int):
    """Set the decoder layer count where it actually lives in the HF config.

    Module-level (picklable) so it can be passed as a ``hf_overrides`` callable
    to vLLM, which pickles the config when spawning the EngineCore subprocess.
    Some multimodal models (e.g. MiniMax-M3) nest ``num_hidden_layers`` under
    ``text_config``; a top-level override is ignored there.
    """
    text_cfg = getattr(hf_config, "text_config", None)
    if text_cfg is not None and hasattr(text_cfg, "num_hidden_layers"):
        text_cfg.num_hidden_layers = n
    else:
        hf_config.num_hidden_layers = n
    return hf_config


def _build_layer_override(
    profiled_layers: int,
    quantization: str | None,
    layers_under_text_config: bool,
):
    """Build the ``hf_overrides`` value for reduced-layer profiling.

    Normally we return a callable that sets ``num_hidden_layers`` where it
    actually lives (top level, or nested under ``text_config`` for multimodal
    models like MiniMax-M3). vLLM applies callables in place, preserving the
    rest of the config.

    When ``quantization`` is requested, vLLM's ``get_quant_config`` rejects a
    callable ("hf_overrides must be a dict ...") because it reads the quant
    config out of ``hf_overrides``. In that case we return a **dict** override
    targeting the right key instead — vLLM applies nested ``text_config`` dicts
    recursively.
    """
    if quantization:
        if layers_under_text_config:
            return {"text_config": {"num_hidden_layers": profiled_layers}}
        return {"num_hidden_layers": profiled_layers}
    # Module-level partial so it pickles for the spawned EngineCore subprocess.
    return functools.partial(_set_num_hidden_layers, n=profiled_layers)


def _make_token_ids(n: int, vocab_size: int, seed: int) -> list[int]:
    """Deterministically build ``n`` valid, non-special token ids.

    Ids are drawn from a safe interior range of the vocabulary (avoiding the
    low ids that are typically special/control tokens) using a cheap hash of the
    position and ``seed``. Two calls with the same ``(n, vocab_size, seed)``
    yield the identical sequence, which is what lets the prefix-cache warm pass
    and the profiled pass share an exact-match context prefix.
    """
    if n <= 0:
        return []
    lo = 256
    hi = max(lo + 1, vocab_size - 256)
    span = hi - lo
    return [lo + ((i * 2654435761 + seed * 40503 + 12345) % span) for i in range(n)]


def _get_block_size(llm, default: int = 16) -> int:
    """Read the KV-cache block size from a constructed vLLM engine (robust)."""
    engine = getattr(llm, "llm_engine", None)
    for attr in ("vllm_config", "engine_config", "model_config"):
        cfg = getattr(engine, attr, None)
        cc = getattr(cfg, "cache_config", None)
        bs = getattr(cc, "block_size", None)
        if bs:
            return int(bs)
    cc = getattr(engine, "cache_config", None)
    bs = getattr(cc, "block_size", None)
    if bs:
        return int(bs)
    return default


def _get_vocab_size(llm, summary: dict, default: int = 32000) -> int:
    """Best-effort tokenizer vocabulary size for synthetic-prompt generation."""
    try:
        tok = llm.get_tokenizer()
        vs = getattr(tok, "vocab_size", None) or len(tok)
        if vs and vs > 512:
            return int(vs)
    except Exception:
        pass
    vs = summary.get("vocab_size")
    return int(vs) if vs and vs > 512 else default


def _collect_node_names(node: dict | None) -> set[str]:
    """Set of all display ``name`` values in a graph tree (for diagnostics)."""
    if not node:
        return set()
    out = {node.get("name", "")}
    for c in node.get("children", []):
        out |= _collect_node_names(c)
    return out


def _collect_module_types(node: dict | None) -> set[str]:
    """Set of all ``module_type`` (class) values in a graph tree."""
    if not node:
        return set()
    out = {node.get("module_type", "")}
    for c in node.get("children", []):
        out |= _collect_module_types(c)
    return out


def _build_result_from_traces(
    rank_files: list[str],
    *,
    model_id: str,
    summary: dict,
    dim_symbols: dict,
    tp_size: int,
    batch_size: int,
    mode: str = "eager",
    max_model_len: int | None = None,
    max_tokens: int | None = None,
    quantization: str | None = None,
    profiled_layers: int | None = None,
    actual_layers: int | None = None,
    layer_scale: float = 1.0,
    trace_file: str | None = None,
    ref_module_tree: dict | None = None,
    query_len: int | None = None,
    context_len: int | None = None,
) -> dict:
    """Parse one or more trace files and build the profile result dict.

    Shared by the live profiler (``_run_profile``) and the trace-upload
    endpoint so both paths reconstruct the model graph and op breakdown the
    same way. ``rank_files`` is rank-0 first; with TP>1 the remaining ranks are
    used only to average device time.
    """
    from breakdown.trace_parser import parse_trace_file

    op_dicts = parse_trace_file(rank_files[0])

    # If multi-rank, average device times across ranks
    if tp_size > 1 and len(rank_files) > 1:
        for extra_file in rank_files[1:]:
            extra_ops = parse_trace_file(extra_file)
            extra_timing = {
                (o["name"], o.get("input_shapes", "")): o.get("device_time_us", 0)
                for o in extra_ops
            }
            for op in op_dicts:
                key = (op["name"], op.get("input_shapes", ""))
                op["device_time_us"] = (
                    op.get("device_time_us", 0) + extra_timing.get(key, 0)
                )
        for op in op_dicts:
            op["device_time_us"] = op.get("device_time_us", 0) / tp_size

    if not op_dicts:
        raise RuntimeError(
            f"No ops found in trace file {rank_files[0]}. "
            "The trace may not contain any captured events."
        )

    analyzed = analyze_ops(
        op_dicts,
        dim_symbols=dim_symbols,
        batch_size=batch_size,
        seq_len=None,
        model_dtype=summary.get("dtype", "bfloat16"),
        num_layers=summary.get("num_layers"),
    )

    backend_totals: dict[str, dict] = {}
    total_dev = sum(o.device_time_us for o in analyzed)
    for b in Backend:
        ops = [o for o in analyzed if o.backend == b.value]
        dev = sum(o.device_time_us for o in ops)
        backend_totals[b.value] = {
            "device_time_us": dev,
            "pct": round(dev / total_dev * 100, 1) if total_dev > 0 else 0,
            "num_ops": len(ops),
            "num_calls": sum(o.call_count for o in ops),
        }

    profile_result = {
        "model_id": model_id,
        "mode": mode,
        "batch_size": batch_size,
        "max_model_len": max_model_len,
        "max_tokens": max_tokens,
        "tp_size": tp_size,
        "quantization": quantization,
        "summary": summary,
        "total_device_time_us": total_dev,
        "total_cpu_time_us": sum(o.cpu_time_us for o in analyzed),
        "backends": backend_totals,
        "ops": [o.to_dict() for o in analyzed],
        "profiled_layers": profiled_layers,
        "actual_layers": actual_layers,
        "layer_scale": layer_scale,
        "trace_file": trace_file if trace_file is not None else rank_files[0],
    }

    # Reconstruct the model graph directly from the profiler trace.
    try:
        graph = build_graph_from_trace(
            rank_files[0],
            summary=summary,
            tp_size=tp_size,
            batch_size=batch_size,
            quantization=quantization,
            ref_module_tree=ref_module_tree,
            query_len=query_len,
            context_len=context_len,
        )
        graph["profiled_layers"] = profiled_layers
        graph["actual_layers"] = actual_layers
        graph["layer_scale"] = layer_scale
        profile_result["graph"] = graph
        if ref_module_tree:
            names = _collect_node_names(graph.get("prefill")
                                        or graph.get("decode"))
            recovered = names & {"q_norm", "k_norm", "input_layernorm",
                                 "post_attention_layernorm"}
            if recovered:
                logger.info("Module-name recovery applied to graph: %s",
                            ", ".join(sorted(recovered)))
            else:
                logger.warning(
                    "Module-name recovery: reference tree was available but no "
                    "attribute names landed on the graph (structural alignment "
                    "found no match). Trace module classes seen: %s",
                    ", ".join(sorted(_collect_module_types(
                        graph.get("prefill") or graph.get("decode")))[:20]),
                )
    except Exception:
        logger.warning("Graph reconstruction failed", exc_info=True)

    return profile_result


def _merge_two_pass_result(pre: dict, dec: dict,
                           prefill_bs: int, decode_bs: int) -> dict:
    """Splice a prefill-batch pass and a decode-batch pass into one result.

    Real serving decouples the phases: prefill typically runs ~1 sequence at a
    time while decode batches many concurrent sequences. A single
    ``llm.generate`` call cannot express that (it prefills and decodes the same
    batch), so we profile two passes and merge them here:

    - ``pre`` — full result from a pass run at ``prefill_bs`` (its **prefill**
      phase is the faithful one; ``S`` = query_len).
    - ``dec`` — full result from a pass run at ``decode_bs`` (its **decode**
      phase is faithful; ``B`` = decode_bs).

    The merged result keeps the decode pass as the base (its op breakdown
    reflects the steady-state, throughput-bound decode batch) and overlays the
    prefill pass's prefill graph tree, so the reconstructed graph shows
    prefill@``prefill_bs`` together with decode@``decode_bs``.
    """
    result = dict(dec)
    result["batch_size"] = decode_bs
    result["prefill_batch_size"] = prefill_bs
    result["decode_batch_size"] = decode_bs
    result["two_pass"] = True

    gpre = pre.get("graph") or {}
    gdec = dec.get("graph") or {}
    graph = dict(gdec)
    graph["prefill"] = gpre.get("prefill")
    # Symbols: the decode pass supplies ``B`` (decode batch); the prefill pass
    # supplies the prefill token dims ``S`` / ``S+C`` / ``C``.
    sym = dict(gdec.get("symbols") or {})
    presym = gpre.get("symbols") or {}
    for k in ("S", "S+C", "C"):
        if k in presym:
            sym[k] = presym[k]
    graph["symbols"] = sym
    result["graph"] = graph
    return result


def _run_profile(model_id: str, mode: str, max_model_len: int,
                 batch_size: int, max_tokens: int, prompt: str,
                 num_profile_layers: int | None = None,
                 tp_size: int = 1,
                 quantization: str | None = None,
                 gpu_memory_utilization: float | None = None,
                 query_len: int | None = None,
                 context_len: int | None = None,
                 prefill_batch_size: int | None = None,
                 decode_batch_size: int | None = None):
    """Run profiling in a background thread using vLLM's native profiler.

    On XPU hardware, vLLM automatically selects XPUWorker which uses
    the correct profiler activities (["CPU", "XPU"]).

    Args:
        num_profile_layers: If set (e.g. 1), override the model to load only
            this many layers. Timing is then scaled by actual_layers/profiled
            for the full model estimate. Enables profiling models too large
            for the GPU.
        tp_size: tensor parallel size (default 1). With TP>1, vLLM creates
            one trace file per rank; we parse all and aggregate timing.
        quantization: quantization method (e.g. "fp8", "gptq", "awq").
            Passed as --quantization to vLLM.
        gpu_memory_utilization: fraction of device memory vLLM may use. Lower
            it (e.g. 0.8) when vLLM's init footprint leaves too little headroom
            for the default (0.92) on small-VRAM cards. None keeps vLLM default.
        query_len: number of *new* prompt tokens the profiled prefill computes
            (the "Query Len"). Drives the prefill token dimension ``S``.
        context_len: number of prior context tokens the query attends to (the
            "Context Len"). When >0, those tokens are pre-computed in an
            un-profiled warm pass and served from the prefix cache (APC) during
            the profiled run, so the profiled prefill computes only ``query_len``
            new tokens while attention still reads the full ``context_len+query_len``
            KV. Rounded down to a KV block boundary so the whole context caches.
        prefill_batch_size: number of concurrent sequences for the **prefill**
            phase (typically 1 in real serving). When it differs from
            ``decode_batch_size`` the run is profiled in two passes — a prefill
            pass at this batch and a decode pass at ``decode_batch_size`` — and
            the two phase graphs are merged. ``None`` falls back to
            ``batch_size`` (single pass, legacy behaviour).
        decode_batch_size: number of concurrent sequences for the **decode**
            phase (often 32/64/128). See ``prefill_batch_size``. ``None`` falls
            back to ``batch_size``.
    """
    global _profile_state
    try:
        from vllm import LLM, SamplingParams, TokensPrompt

        # Fetch model config for analysis
        try:
            config = fetch_model_config(model_id)
            summary = summarize_config(config)
            dim_symbols = get_dim_symbols(summary)
        except Exception:
            config = {}
            summary = {}
            dim_symbols = {}

        actual_layers = summary.get("num_layers") or 1
        if num_profile_layers == "min":
            # Auto-calculate minimum layers needed
            profiled_layers = min_profile_layers(summary)
        elif num_profile_layers:
            profiled_layers = int(num_profile_layers)
        else:
            profiled_layers = actual_layers
        layer_scale = actual_layers / profiled_layers

        trace_dir = os.path.abspath("output/traces")
        os.makedirs(trace_dir, exist_ok=True)

        engine_kwargs: dict = {
            "model": model_id,
            "max_model_len": max_model_len,
            "tensor_parallel_size": tp_size,
            "profiler_config": {
                "profiler": "torch",
                "torch_profiler_dir": trace_dir,
                "torch_profiler_record_shapes": True,
                "torch_profiler_with_stack": True,
                "torch_profiler_with_flops": True,
                "torch_profiler_use_gzip": True,
            },
        }

        # Always use dummy weights for profiling — timing doesn't depend on
        # weight values, and dummy avoids KeyError when layers are reduced.
        engine_kwargs["load_format"] = "dummy"

        # Optionally cap device memory usage (leaves headroom for vLLM's init
        # footprint on small-VRAM cards; None keeps vLLM's default).
        if gpu_memory_utilization is not None:
            engine_kwargs["gpu_memory_utilization"] = gpu_memory_utilization

        # Vision-language models: disable multimodal memory profiling so the run
        # captures the language-model ops on a text prompt. This avoids vLLM's
        # dummy image/video profiling path through the vision tower (which the
        # static graph already covers) and keeps the profile focused on the LLM.
        if summary.get("vit_hidden_size"):
            engine_kwargs["limit_mm_per_prompt"] = {"image": 0, "video": 0}

        # Sparse-attention models (e.g. MiniMax-M3) select fixed-size KV blocks
        # via the lightning indexer, so the KV-cache block size must match the
        # sparse block size; otherwise vLLM cannot reconcile a common kernel
        # block size across the sparse/full attention backends.
        sparse_block = summary.get("sparse_block_size")
        if sparse_block:
            engine_kwargs["block_size"] = int(sparse_block)

        # Normalize the query/context sizing knobs. ``query_len`` sets the
        # number of new prompt tokens the profiled prefill computes; when
        # ``context_len`` > 0 we serve that many prior tokens from the prefix
        # cache so the profiled prefill sees ``S = query_len`` new tokens
        # attending to a ``context_len``-token KV context.
        query_len = int(query_len) if query_len else 0
        context_len = int(context_len) if context_len else 0
        use_token_prompts = query_len > 0
        # The context length actually served from the prefix cache, floored to a
        # whole number of KV blocks (set below once the block size is known). The
        # graph reconstruction symbolizes this value as ``C`` so attention KV
        # dims read ``C`` / ``S+C`` instead of a bare number.
        profiled_context_len = 0

        # Enable Automatic Prefix Caching so the context prefix computed in the
        # warm pass is reused (not recomputed) during the profiled run.
        if context_len > 0:
            engine_kwargs["enable_prefix_caching"] = True

        # Quantization method
        if quantization:
            engine_kwargs["quantization"] = quantization

        # Override layer count for reduced-layer profiling.
        if profiled_layers < actual_layers:
            # Some multimodal models (e.g. MiniMax-M3) nest the decoder layer
            # count under ``text_config``. A top-level ``num_hidden_layers``
            # override is silently ignored there, so the full model is built
            # and exhausts device memory (UR_RESULT_ERROR_DEVICE_LOST). The
            # helper returns a callable normally, but a dict when quantization
            # is set (vLLM's ``get_quant_config`` requires a dict override).
            engine_kwargs["hf_overrides"] = _build_layer_override(
                profiled_layers,
                quantization,
                bool(summary.get("layers_under_text_config")),
            )

        # Set compile / eager mode
        if mode == "compile":
            os.environ["VLLM_TORCH_COMPILE_LEVEL"] = "3"
            engine_kwargs["enforce_eager"] = False
        else:
            os.environ.pop("VLLM_TORCH_COMPILE_LEVEL", None)
            engine_kwargs["enforce_eager"] = True

        # Resolve the per-phase batch sizes. When prefill and decode batches
        # differ we profile two passes (real serving prefills ~1 sequence while
        # decoding many); otherwise a single pass reproduces legacy behaviour.
        pf_batch = int(prefill_batch_size) if prefill_batch_size else int(batch_size)
        dc_batch = int(decode_batch_size) if decode_batch_size else int(batch_size)
        two_pass = pf_batch != dc_batch
        max_batch = max(pf_batch, dc_batch)

        # Pin the scheduler so every decode step runs the *full* requested batch
        # (``B = decode_batch``). Left to its defaults, vLLM's continuous-batching
        # scheduler caps per-iteration concurrency (by ``max_num_seqs`` and by how
        # many sequences' KV fits in cache) and runs an oversized batch in
        # *partial-batch waves* — e.g. a batch of 32 dispatched as 29 + 3. Each
        # wave has a different row count, so its ops neither symbolize to ``B``
        # nor merge with the full-batch ops, surfacing as duplicated ``29``/``3``
        # nodes in the reconstructed decode graph. ``max_num_seqs = max_batch``
        # forces the scheduler to admit the whole batch in one iteration;
        # ``max_num_batched_tokens`` is sized to also admit a whole batch's
        # prefill tokens in a single step (prefill pass: ``pf_batch × query_len``;
        # decode pass: ``dc_batch`` single-token prefills) so a full-shape step is
        # never chunked. If the batch's KV cannot fit device memory, raise
        # ``gpu_memory_utilization`` or lower Context/Batch rather than letting the
        # run silently split.
        engine_kwargs["max_num_seqs"] = max_batch
        _prefill_step_tokens = pf_batch * max(int(query_len), 1)
        engine_kwargs["max_num_batched_tokens"] = max(
            _prefill_step_tokens, max_batch, 2048)

        llm = LLM(**engine_kwargs)

        if use_token_prompts:
            block_size = _get_block_size(llm)
            vocab_size = _get_vocab_size(llm, summary)
            # Round the context down to a whole number of KV blocks so the entire
            # context prefix is cacheable (a trailing partial block would be
            # recomputed and shift the profiled prefill token count).
            ctx_aligned = (context_len // block_size) * block_size
            profiled_context_len = ctx_aligned
            ctx_ids = _make_token_ids(ctx_aligned, vocab_size, seed=0)
        else:
            ctx_aligned = 0
            ctx_ids = []

        def _list_trace_files() -> set:
            return {os.path.join(trace_dir, f) for f in os.listdir(trace_dir)
                    if f.endswith(".json") or f.endswith(".json.gz")}

        def _install_span_hooks() -> bool:
            """Install capture-time module-name span hooks in the worker(s).

            Registers forward hooks that emit ``record_function(
            "module::<qname>::<Cls>")`` spans around every module's forward, so
            the trace carries real attribute names (``q_norm``/``k_norm``,
            ``self_attn``, ...) and ``build_graph_from_trace`` reconstructs the
            tree with exact names — no reference-tree overlay needed (research
            R1). Best-effort: on any failure the run proceeds and naming falls
            back to the ``module_naming`` overlay.
            """
            try:
                from breakdown.module_hooks import install_module_span_hooks_on
                llm.apply_model(install_module_span_hooks_on)
                return True
            except Exception:
                logger.warning("module span hooks: install failed; module names "
                               "will fall back to the name overlay",
                               exc_info=True)
                return False

        def _remove_span_hooks(installed: bool) -> None:
            if not installed:
                return
            try:
                from breakdown.module_hooks import remove_module_span_hooks_on
                llm.apply_model(remove_module_span_hooks_on)
            except Exception:
                logger.warning("module span hooks: remove failed", exc_info=True)

        def _profiled_pass(pass_batch: int, pass_query_len: int,
                           pass_max_tokens: int):
            """Warm + run one profiled generate at the given batch/query size.

            ``pass_max_tokens`` is the number of tokens to generate this pass:
            the decode pass uses the full decode budget (so decode steps are
            captured), while the prefill pass uses **1** — it only needs the
            single prefill step (``S`` = ``query_len``), and generating extra
            decode tokens would only bloat the trace and slow the run.

            ``ignore_eos`` keeps every sequence alive for the full budget so the
            profiled decode step reflects the requested batch (a sequence hitting
            EOS early would shrink the observed decode concurrency ``B``).

            Returns ``(rank_files, cache_hit_note)``: the trace file(s) newly
            written by this pass (newest first, capped to ``tp_size``) and an
            optional prefix-cache-miss note.
            """
            note = None
            pass_sampling = SamplingParams(max_tokens=pass_max_tokens,
                                           ignore_eos=True)
            if use_token_prompts:
                def _full_prompt(query_seed: int) -> "TokensPrompt":
                    q = _make_token_ids(pass_query_len, vocab_size, seed=query_seed)
                    return TokensPrompt(prompt_token_ids=ctx_ids + q)

                # Distinct query per batch item (shared context prefix) so every
                # sequence genuinely prefills its own tokens instead of
                # cache-hitting a sibling; profiled seeds differ from warmup
                # seeds so the profiled queries are never served from cache.
                def _batch(base_seed: int) -> list["TokensPrompt"]:
                    return [_full_prompt(base_seed + b) for b in range(pass_batch)]

                # Warm the shared prefix cache once (un-profiled) so the profiled
                # run reads the context from cache instead of recomputing it.
                if ctx_ids:
                    llm.generate(
                        [TokensPrompt(prompt_token_ids=ctx_ids)],
                        SamplingParams(max_tokens=1), use_tqdm=False,
                    )
                for w in range(2):
                    llm.generate(_batch(1000 * (w + 1)), pass_sampling,
                                 use_tqdm=False)
                profiled_prompts = _batch(900000)

                before = _list_trace_files()
                _spans = _install_span_hooks()
                llm.start_profile()
                outputs = llm.generate(profiled_prompts, pass_sampling,
                                       use_tqdm=False)
                llm.stop_profile()
                _remove_span_hooks(_spans)

                # Verify the context was served from cache. A miss means the
                # profiled prefill recomputed the whole context (S = context +
                # query), so record a note rather than silently misreporting.
                if context_len > 0 and outputs:
                    cached = getattr(outputs[0], "num_cached_tokens", None)
                    if cached is not None and cached < ctx_aligned:
                        note = (
                            f"Prefix cache hit only {cached}/{ctx_aligned} "
                            "context tokens; profiled prefill may include "
                            "context recompute."
                        )
            else:
                # --- Legacy text-prompt path (Query/Context Len unspecified) ---
                conversation = [
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": prompt},
                ]
                conversations = [conversation] * pass_batch
                prompts = [prompt] * pass_batch

                # First warmup pass also detects chat-template support.
                try:
                    llm.chat(conversations, pass_sampling, use_tqdm=False)
                    use_chat = True
                except Exception:
                    llm.generate(prompts, pass_sampling, use_tqdm=False)
                    use_chat = False

                def _run_inference():
                    if use_chat:
                        llm.chat(conversations, pass_sampling, use_tqdm=False)
                    else:
                        llm.generate(prompts, pass_sampling, use_tqdm=False)

                # Warmup primes Triton JIT / autotuning so the profiled trace
                # reflects steady-state timing. Detection pass above is warmup
                # #1; run 2 more for 3 total.
                for _ in range(2):
                    _run_inference()

                before = _list_trace_files()
                _spans = _install_span_hooks()
                llm.start_profile()
                _run_inference()
                llm.stop_profile()
                _remove_span_hooks(_spans)

            # torch's profiler writes the trace on stop_profile; wait briefly for
            # the new file(s) to appear, then return them (newest first).
            new_files: list[str] = []
            for _ in range(20):
                new_files = sorted(_list_trace_files() - before,
                                   key=os.path.getmtime, reverse=True)
                if len(new_files) >= tp_size:
                    break
                time.sleep(0.5)
            if not new_files:
                new_files = sorted(_list_trace_files(),
                                   key=os.path.getmtime, reverse=True)
            return new_files[:tp_size], note

        # --- Run the profiled pass(es) ---
        if two_pass:
            # Prefill pass: batch = prefill_batch, real query_len, generate only
            # 1 token so the trace holds exactly the prefill step (S=query_len)
            # — we keep only its prefill phase. Decode pass: batch = decode_batch,
            # query_len forced to 1 so decode is 1 new token/seq (matches real
            # decode and avoids OOM from prefilling decode_batch x query_len
            # tokens), generating the full decode budget — we keep its decode
            # phase.
            pre_files, note_pre = _profiled_pass(pf_batch, query_len,
                                                 pass_max_tokens=1)
            dec_query = 1 if use_token_prompts else 0
            dec_files, note_dec = _profiled_pass(dc_batch, dec_query,
                                                 pass_max_tokens=max_tokens)
            cache_hit_note = note_pre or note_dec
        else:
            # Single pass yields both phases from one run, so it needs the full
            # decode budget.
            single_files, cache_hit_note = _profiled_pass(
                dc_batch, query_len, pass_max_tokens=max_tokens)

        # Module attribute names (q_norm/k_norm, ...) come primarily from the
        # capture-time ``module::`` spans installed above, which the
        # reconstruction reads directly from the trace. We still capture a
        # ``named_modules()`` reference tree from the live model as a *fallback*
        # overlay: ``build_graph_from_trace`` uses it only if the trace lacks
        # those spans (e.g. hook install failed, or an older trace).
        ref_module_tree = None
        try:
            from breakdown.module_naming import ref_tree_from_llm
            ref_module_tree = ref_tree_from_llm(llm)
        except Exception:
            logger.warning("ref_tree_from_llm raised; fallback name overlay "
                           "disabled", exc_info=True)
            ref_module_tree = None

        if ref_module_tree:
            logger.info(
                "Module-name overlay (fallback): got reference tree from live "
                "model (root=%s, %d top-level children)",
                ref_module_tree.get("cls"),
                len(ref_module_tree.get("children", [])),
            )
        else:
            logger.warning(
                "Module-name overlay (fallback): could NOT read module names "
                "from the live model (ref_tree_from_llm returned None). If the "
                "capture-time spans also failed, q_norm/k_norm and other "
                "attribute names will fall back to class heuristics."
            )
            # Fallback: the offline meta-device path. Heavy (re-instantiates on
            # meta) but lets naming work when the live-model traversal fails.
            if config:
                try:
                    from breakdown.module_naming import ref_tree_from_config
                    ref_module_tree = ref_tree_from_config(
                        config,
                        dtype=summary.get("dtype", "bfloat16"),
                        model_id=model_id,
                    )
                    if ref_module_tree:
                        logger.info(
                            "Module-name overlay (fallback): recovered names via "
                            "meta-device fallback."
                        )
                except Exception:
                    logger.warning("ref_tree_from_config fallback failed",
                                   exc_info=True)
                    ref_module_tree = None

        # --- Parse trace files & build the result ---
        # With TP>1, vLLM produces one trace file per rank; each pass's
        # ``_profiled_pass`` already returned that pass's rank-0-first files.
        def _build(files: list[str], bsz: int, qlen: int | None) -> dict:
            if not files:
                raise RuntimeError(
                    f"No trace files found in {trace_dir}. "
                    "Profiling may have failed in the worker process."
                )
            return _build_result_from_traces(
                files,
                model_id=model_id,
                summary=summary,
                dim_symbols=dim_symbols,
                tp_size=tp_size,
                batch_size=bsz,
                mode=mode,
                max_model_len=max_model_len,
                max_tokens=max_tokens,
                quantization=quantization,
                profiled_layers=profiled_layers,
                actual_layers=actual_layers,
                layer_scale=layer_scale,
                ref_module_tree=ref_module_tree,
                query_len=qlen,
                context_len=profiled_context_len or None,
            )

        if two_pass:
            res_pre = _build(pre_files, pf_batch, query_len or None)
            res_dec = _build(dec_files, dc_batch,
                             1 if use_token_prompts else None)
            profile_result = _merge_two_pass_result(
                res_pre, res_dec, pf_batch, dc_batch)
        else:
            profile_result = _build(single_files, dc_batch, query_len or None)

        profile_result["query_len"] = query_len or None
        profile_result["context_len"] = context_len or None
        profile_result["context_len_aligned"] = profiled_context_len or None
        if cache_hit_note:
            profile_result["cache_hit_note"] = cache_hit_note

        with _profile_lock:
            _profile_state["status"] = "done"
            _profile_state["result"] = profile_result
            _profile_state["error"] = None

    except Exception as e:
        with _profile_lock:
            _profile_state["status"] = "error"
            _profile_state["error"] = traceback.format_exc()
    finally:
        os.environ.pop("VLLM_TORCH_COMPILE_LEVEL", None)
        try:
            import torch.distributed as dist
            if dist.is_initialized():
                dist.destroy_process_group()
        except Exception:
            pass


@app.route("/api/profile", methods=["POST"])
def start_profile():
    """Start a profiling run. Non-blocking — poll /api/profile/status."""
    global _profile_state

    with _profile_lock:
        if _profile_state["status"] == "running":
            return jsonify({"ok": False, "error": "Profiling already in progress"}), 409

    data = request.json or {}
    model_id = data.get("model_id", "")
    if not model_id:
        return jsonify({"ok": False, "error": "model_id is required"}), 400

    mode = data.get("mode", "eager")
    max_model_len = data.get("max_model_len", 4096)
    batch_size = data.get("batch_size", 1)
    max_tokens = data.get("max_tokens", 128)
    prompt = data.get("prompt", "Write a short essay about AI.")
    num_profile_layers = data.get("num_profile_layers")  # None = all layers
    tp_size = data.get("tensor_parallel_size", 1)
    quantization = data.get("quantization")  # None = no quantization
    gpu_memory_utilization = data.get("gpu_memory_utilization")  # None = vLLM default
    query_len = data.get("query_len")  # new prefill tokens (None = legacy prompt)
    context_len = data.get("context_len")  # cached prefix tokens (None/0 = none)
    # Per-phase batch sizes. When they differ, profiling runs two passes
    # (prefill@prefill_batch_size + decode@decode_batch_size) and merges them.
    # Absent → fall back to the single ``batch_size`` (legacy single pass).
    prefill_batch_size = data.get("prefill_batch_size")
    decode_batch_size = data.get("decode_batch_size")

    # The engine must fit the whole sequence it will ever see: cached context +
    # new query tokens + the decode tokens we generate. The frontend sizes
    # max_model_len from Query+Context; bump it to also cover the decode budget.
    if query_len:
        needed = int(query_len) + int(context_len or 0) + int(max_tokens) + 16
        if needed > int(max_model_len):
            max_model_len = needed

    with _profile_lock:
        _profile_state = {
            "status": "running",
            "result": None,
            "error": None,
            "model_id": model_id,
            "settings": {
                "mode": mode,
                "batch_size": batch_size,
                "prefill_batch_size": prefill_batch_size,
                "decode_batch_size": decode_batch_size,
                "max_model_len": max_model_len,
                "tp_size": tp_size,
                "quantization": quantization,
                "query_len": query_len,
                "context_len": context_len,
            },
        }

    thread = threading.Thread(
        target=_run_profile,
        args=(model_id, mode, max_model_len, batch_size, max_tokens, prompt,
              num_profile_layers, tp_size, quantization, gpu_memory_utilization,
              query_len, context_len, prefill_batch_size, decode_batch_size),
        daemon=True,
    )
    thread.start()

    return jsonify({"ok": True, "status": "running"})


@app.route("/api/profile/upload", methods=["POST"])
def upload_profile():
    """Reconstruct the model graph and op breakdown from uploaded trace(s).

    Accepts a multipart form with one or more ``trace`` files (a torch profiler
    Chrome trace, ``.json`` or ``.json.gz``; with TP>1 upload one file per rank,
    rank-0 first) plus optional form fields:

      - ``model_id``: HF id used to fetch config for shape symbols / summary
      - ``tensor_parallel_size`` / ``tp_size``: ranks represented by the uploads
      - ``batch_size``, ``quantization``, ``mode``
      - ``num_profile_layers`` / ``actual_layers``: for reduced-layer scaling

    Parsing is fast, so this runs synchronously and stores the result in the
    shared profile state so ``/api/profile/result`` and ``/api/profile/trace``
    work exactly as they do for a live profiling run.
    """
    global _profile_state

    with _profile_lock:
        if _profile_state["status"] == "running":
            return jsonify({"ok": False, "error": "Profiling already in progress"}), 409

    files = request.files.getlist("trace")
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({"ok": False, "error": "No trace file uploaded"}), 400

    form = request.form
    model_id = (form.get("model_id") or "").strip()
    mode = form.get("mode", "eager")
    tp_size = int(form.get("tensor_parallel_size") or form.get("tp_size") or 1)
    batch_size = int(form.get("batch_size") or 1)
    quantization = form.get("quantization") or None
    if quantization in ("", "auto", "none"):
        quantization = None

    # Persist uploads under output/traces so the trace-download endpoint works.
    from werkzeug.utils import secure_filename
    trace_dir = os.path.abspath("output/traces")
    os.makedirs(trace_dir, exist_ok=True)
    saved: list[str] = []
    for f in files:
        name = secure_filename(f.filename) or "uploaded_trace.json"
        dest = os.path.join(trace_dir, name)
        f.save(dest)
        saved.append(dest)

    # Fetch model config for shape symbols / summary (best-effort).
    try:
        summary = summarize_config(fetch_model_config(model_id)) if model_id else {}
        dim_symbols = get_dim_symbols(summary) if summary else {}
    except Exception:
        summary = {}
        dim_symbols = {}

    actual_layers = form.get("actual_layers") or summary.get("num_layers")
    actual_layers = int(actual_layers) if actual_layers else None
    profiled_layers = form.get("num_profile_layers") or actual_layers
    profiled_layers = int(profiled_layers) if profiled_layers else None
    layer_scale = (
        actual_layers / profiled_layers
        if actual_layers and profiled_layers else 1.0
    )

    with _profile_lock:
        _profile_state = {
            "status": "running",
            "result": None,
            "error": None,
            "model_id": model_id,
            "settings": {
                "mode": mode,
                "batch_size": batch_size,
                "tp_size": tp_size,
                "quantization": quantization,
                "uploaded": True,
            },
        }

    # Reconstruct real module attribute names by instantiating the model on
    # ``meta`` (no weights). Heavy + network-dependent. Without a recovered
    # tree, uploaded traces keep heuristic names.
    ref_module_tree = None
    if model_id:
        try:
            from breakdown.module_naming import ref_tree_from_config
            ref_module_tree = ref_tree_from_config(
                fetch_model_config(model_id),
                dtype=summary.get("dtype", "bfloat16"),
                model_id=model_id,
            )
        except Exception:
            ref_module_tree = None

    try:
        result = _build_result_from_traces(
            saved[:tp_size] if len(saved) >= tp_size else saved,
            model_id=model_id,
            summary=summary,
            dim_symbols=dim_symbols,
            tp_size=tp_size,
            batch_size=batch_size,
            mode=mode,
            quantization=quantization,
            profiled_layers=profiled_layers,
            actual_layers=actual_layers,
            layer_scale=layer_scale,
            ref_module_tree=ref_module_tree,
        )
        with _profile_lock:
            _profile_state["status"] = "done"
            _profile_state["result"] = result
            _profile_state["error"] = None
    except Exception:
        err = traceback.format_exc()
        with _profile_lock:
            _profile_state["status"] = "error"
            _profile_state["error"] = err
        return jsonify({"ok": False, "error": err}), 500

    return jsonify({"ok": True, "status": "done"})


@app.route("/api/profile/status")
def profile_status():
    """Check profiling status."""
    with _profile_lock:
        return jsonify({
            "status": _profile_state["status"],
            "model_id": _profile_state["model_id"],
            "settings": _profile_state.get("settings"),
            "error": _profile_state["error"],
        })


@app.route("/api/profile/result")
def profile_result():
    """Get profiling results (only available when status=done)."""
    with _profile_lock:
        if _profile_state["status"] != "done":
            return jsonify({
                "ok": False,
                "status": _profile_state["status"],
                "error": _profile_state.get("error"),
            }), 202
        result = _profile_state["result"]
    # Don't expose internal trace_file path; indicate availability
    client_result = {k: v for k, v in result.items() if k != "trace_file"}
    client_result["has_trace"] = bool(result.get("trace_file"))
    return jsonify({"ok": True, "data": client_result})


@app.route("/api/profile/trace")
def download_trace():
    """Download the profiled trace file with a descriptive filename."""
    with _profile_lock:
        if _profile_state["status"] != "done" or not _profile_state.get("result"):
            return jsonify({"ok": False, "error": "No profile result available"}), 404
        result = _profile_state["result"]

    trace_path = result.get("trace_file")
    if not trace_path or not os.path.isfile(trace_path):
        return jsonify({"ok": False, "error": "Trace file not found"}), 404

    # Build a descriptive filename encoding the profiled configuration:
    #   vllm_trace_{model}_{mode}_ctx{context}_in{query}_out{gen}_bs{bs}_tp{tp}_{n}layers.json.gz
    # where "ctx" is the block-aligned prefix-cache context the prefill attends
    # to, "in" is the query length (new prefill tokens, S), "out" is the number
    # of generated decode tokens, and "bs" is the decode batch. The model id is
    # reduced to its final path component (org prefix dropped).
    model_short = result["model_id"].split("/")[-1]
    mode = result.get("mode", "eager")
    bs = result.get("decode_batch_size", result.get("batch_size", 1))
    ctx = result.get("context_len_aligned") or result.get("context_len") or 0
    qin = result.get("query_len") or 0
    gen = result.get("max_tokens", "")
    tp = result.get("tp_size", 1) or 1
    quant = result.get("quantization")
    layers = result.get("profiled_layers", "all")
    device = _DEVICE.upper()
    ext = ".json.gz" if trace_path.endswith(".gz") else ".json"
    quant_part = f"_{quant}" if quant else ""
    download_name = (
        f"vllm_trace_{model_short}_{device}_{mode}_ctx{ctx}_in{qin}_out{gen}"
        f"_bs{bs}_tp{tp}{quant_part}_{layers}layers{ext}"
    )

    return send_file(
        trace_path,
        mimetype="application/gzip" if ext == ".json.gz" else "application/json",
        as_attachment=True,
        download_name=download_name,
    )


# Roles whose 2nd input tensor (index 1) is a weight matrix
_WEIGHT_ROLES = {
    "qkv_proj", "o_proj", "gate_up_proj", "down_proj",
    "expert_gate_up", "expert_down",
    "shared_expert_gate_up", "shared_expert_down",
    "q_compress", "q_decompress", "kv_compress",
    "router_gate", "lm_head", "vl_projector",
    "vit_qkv_proj", "vit_o_proj", "vit_mlp_up", "vit_mlp_down",
    "patch_embed", "mlp_up", "mlp_down",
}


def _bytes_to_dtype(nbytes: int) -> str:
    """Convert byte count to short dtype name."""
    return {4: "fp32", 2: "bf16", 1: "fp8"}.get(nbytes, f"{nbytes * 8}bit")


def _quant_dtype_name(quant: str | None) -> str | None:
    """Map quantization method to weight dtype short name."""
    if not quant or quant == "None":
        return None
    names = {
        "fp8": "fp8", "gptq": "int4", "gptq_marlin": "int4",
        "awq": "int4", "awq_marlin": "int4", "marlin": "int4",
        "squeezellm": "int4", "bitsandbytes": "nf4", "gguf": "q4",
        "int4": "int4", "int8": "int8",
    }
    return names.get(quant.lower(), quant.lower())


def _get_tensor_dtype(tensor_idx: int, role: str,
                      graph_cfg: dict) -> str:
    """Get the dtype label for a specific tensor in an op.

    Mirrors the frontend getOpDtypes logic: weight tensors (index 1) of
    projection ops use the weight dtype; everything else uses activation dtype.
    """
    act_dtype = _bytes_to_dtype(graph_cfg.get("dtype_bytes", 2))
    quant = graph_cfg.get("quantization")
    if quant:
        w_dtype = _quant_dtype_name(quant) or _bytes_to_dtype(
            graph_cfg.get("weight_dtype_bytes", 2)
        )
    else:
        w_dtype = act_dtype

    if tensor_idx == 1 and role in _WEIGHT_ROLES:
        return w_dtype
    return act_dtype


def _flatten_graph_nodes(node: dict, depth: int = 0,
                         rows: list | None = None,
                         parent_repeat: int = 1) -> list[dict]:
    """Flatten a hierarchical graph tree into rows for the hierarchy sheet."""
    if rows is None:
        rows = []

    # Effective repeat = own repeat_count × parent's repeat
    own_repeat = node.get("repeat_count", 1)
    effective_repeat = own_repeat * parent_repeat

    # Add module row
    rows.append({
        "depth": depth,
        "name": node.get("name", ""),
        "path": node.get("path", ""),
        "module_type": node.get("module_type", ""),
        "repeat_count": own_repeat,
        "effective_repeat": effective_repeat,
        "total_memory": node.get("total_memory", 0),
        "total_flops": node.get("total_flops", 0),
        "total_ai": node.get("total_ai", 0),
        "ops": node.get("ops", []),
    })

    # Recurse into children
    for child in node.get("children", []):
        _flatten_graph_nodes(child, depth + 1, rows, effective_repeat)

    return rows


# ---- Shape Matrix Export (single model, multi-config sweep) ----

# Max total rows to prevent excessive memory/time
_MAX_MATRIX_ROWS = 50000


def _format_op_shape_with_dtypes(
    op: dict, symbols: dict[str, int], graph_cfg: dict,
    recorded_dtypes: list[str] | None = None,
) -> str:
    """Format op shapes as concrete values with per-tensor dtypes.

    Example: "[128, 2560, bf16] × [2560, 6144, fp8]"
    Resolves symbolic dims (including composite like "S+C") via the symbols dict.

    When ``recorded_dtypes`` is given (the real per-tensor dtypes captured in the
    profiling trace, aligned with ``op["input_shapes"]``), it takes precedence
    over the config-driven ``_get_tensor_dtype`` heuristic so the exported dtype
    is exactly what the op actually ran.
    """
    op_shapes = op.get("input_shapes", [])
    if not op_shapes:
        return "—"
    role = op.get("role", "")
    parts = []
    for shape_idx, shape in enumerate(op_shapes):
        if isinstance(shape, list):
            tensor_dtype = None
            if recorded_dtypes and shape_idx < len(recorded_dtypes):
                tensor_dtype = _friendly_dtype(recorded_dtypes[shape_idx])
            if not tensor_dtype:
                tensor_dtype = _get_tensor_dtype(shape_idx, role, graph_cfg)
            dims = []
            for dim in shape:
                dims.append(str(_resolve_dim(dim, symbols)))
            if tensor_dtype:
                dims.append(tensor_dtype)
            parts.append("[" + ", ".join(dims) + "]")
        else:
            parts.append(str(_resolve_dim(shape, symbols)))
    return " × ".join(parts)


def _safe_arithmetic_eval(expr: str) -> int:
    """Safely evaluate a simple arithmetic expression (integers, +, -, *, /).

    Only allows integer literals and the operators +, -, *, /.
    Division is performed as integer (floor) division.
    Raises ValueError for anything else.
    """
    import ast

    # Normalize "/" to "//" for integer division in eval
    expr = expr.replace("//", "/").replace("/", "//")

    tree = ast.parse(expr, mode="eval")
    for node in ast.walk(tree):
        if isinstance(node, ast.Expression):
            continue
        if isinstance(node, ast.BinOp):
            if not isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.FloorDiv)):
                raise ValueError(f"Unsupported operator: {type(node.op).__name__}")
        elif isinstance(node, ast.UnaryOp):
            if not isinstance(node.op, (ast.USub, ast.UAdd)):
                raise ValueError(f"Unsupported unary op: {type(node.op).__name__}")
        elif isinstance(node, (ast.Constant,)):
            if not isinstance(node.value, (int, float)):
                raise ValueError(f"Non-numeric constant: {node.value}")
        elif not isinstance(node, (ast.Add, ast.Sub, ast.Mult, ast.FloorDiv,
                                   ast.USub, ast.UAdd)):
            raise ValueError(f"Unsupported node: {type(node).__name__}")
    return int(eval(compile(tree, "<dim>", "eval")))  # noqa: S307


def _resolve_dim(dim, symbols: dict[str, int]):
    """Resolve a dimension value to a concrete integer if possible."""
    if isinstance(dim, int):
        return dim
    if isinstance(dim, str):
        # Direct lookup
        if dim in symbols:
            return symbols[dim]
        # Try evaluating composite expressions like "S+C", "2·I"
        # Replace symbol names with their values and evaluate
        expr = dim
        # Sort by length descending to avoid partial replacements
        for name in sorted(symbols.keys(), key=len, reverse=True):
            expr = expr.replace(name, str(symbols[name]))
        # Replace middle-dot with *
        expr = expr.replace("·", "*")
        try:
            return _safe_arithmetic_eval(expr)
        except (ValueError, SyntaxError, ZeroDivisionError, OverflowError):
            return dim
    return dim


# Config-dependent variable symbols that should stay symbolic
_VARIABLE_SYMS = {"S", "B", "C", "TP"}


def _is_variable_composite(expr: str) -> bool:
    """Check if an expression is composed entirely of variable symbols.

    Handles both additive (S+C) and multiplicative (B·S) composites.
    """
    # Split on + and · to get individual parts
    parts = expr.replace("·", "+").split("+")
    return all(p.strip() in _VARIABLE_SYMS for p in parts)


def _partially_resolve_dim(dim, symbols: dict[str, int],
                           full_symbols: dict[str, int] | None = None,
                           tp_divided: set[str] | None = None):
    """Resolve dim keeping only S/B/C/TP symbolic, resolving all else to numbers.

    Model constants from config.json are shown as numbers. When a dimension
    contains "/TP", it's shown as "value/TP" using the full undivided config
    value from the symbols dict.

    The full_symbols and tp_divided params are accepted for backwards
    compatibility but ignored — the graph now embeds /TP directly in shapes
    and symbols already contain original (undivided) values.
    """
    if isinstance(dim, (int, float)):
        return str(int(dim))

    s = str(dim)

    # Pure variable symbol → keep as-is
    if s in _VARIABLE_SYMS:
        return s

    # Composite of only variable symbols (e.g. "S+C", "B·S") → keep as-is
    if _is_variable_composite(s):
        return s

    # Handle "/TP" suffix: resolve the base part, keep /TP
    if s.endswith("/TP"):
        base = s[:-3]  # strip "/TP"
        resolved_base = _resolve_constant_expr(base, symbols)
        return f"{resolved_base}/TP"

    # Check if s is a known symbol directly (handles names with · like "n_h·D_qh")
    if s in symbols:
        return str(symbols[s])

    # Check for multiply composites containing a variable (e.g., "B·S·K")
    if "·" in s:
        parts = s.split("·")
        has_variable = any(p in _VARIABLE_SYMS for p in parts)
        if has_variable:
            # Partially resolve: keep variable parts, resolve constants
            resolved_parts = []
            for p in parts:
                if p in _VARIABLE_SYMS:
                    resolved_parts.append(p)
                elif p in symbols:
                    resolved_parts.append(str(symbols[p]))
                elif p.isdigit():
                    resolved_parts.append(p)
                else:
                    resolved_parts.append(p)
            return "·".join(resolved_parts)
        else:
            # All parts are constants — compute product
            product = 1
            for p in parts:
                val = symbols.get(p)
                if val is not None:
                    product *= val
                elif p.isdigit():
                    product *= int(p)
            return str(product)

    # Pure constant — fully resolve
    if s in symbols:
        return str(symbols[s])
    resolved = _resolve_dim(s, symbols)
    return str(resolved)


def _resolve_constant_expr(expr: str, symbols: dict[str, int]) -> str:
    """Resolve a constant expression (no variables) to its numeric value.

    Handles symbols like "QKV", "n_h·d", "2·I", and plain numbers.
    """
    if expr in symbols:
        return str(symbols[expr])
    if "·" in expr:
        parts = expr.split("·")
        product = 1
        for p in parts:
            val = symbols.get(p)
            if val is not None:
                product *= val
            elif p.isdigit():
                product *= int(p)
            else:
                return expr  # can't resolve
        return str(product)
    if expr.isdigit():
        return expr
    return expr


# ---- Profile-derived Shape Matrix helpers ----

_FRIENDLY_DTYPE = {
    "bfloat16": "bf16", "bf16": "bf16",
    "float16": "fp16", "fp16": "fp16",
    "float32": "fp32", "float": "fp32", "fp32": "fp32",
    "float8_e4m3fn": "fp8e4m3", "float8_e5m2": "fp8e5m2", "fp8": "fp8",
    "int8": "int8", "uint8": "uint8",
    "int64": "i64", "int32": "i32", "int16": "i16", "long": "i64", "int": "i32",
    "bool": "bool",
}


def _friendly_dtype(name: str) -> str:
    """Short display name for a recorded trace dtype ('bfloat16' → 'bf16')."""
    if not name:
        return ""
    return _FRIENDLY_DTYPE.get(name.lower(), name.lower())


def _prod_ints(shape) -> int:
    p = 1
    for d in shape:
        if isinstance(d, int):
            p *= d
        else:
            return 0
    return p


_MM_OP_BASES = {"mm", "addmm", "linear", "matmul", "bmm", "_scaled_mm",
                "fp8_gemm", "fp4_gemm", "int4_gemm_w4a16", "int4_gemm_w4a8"}


def _profile_op_memory(op_name: str, shapes: list[list[int]],
                       dtypes: list[str], act_bytes: int) -> int:
    """Estimate op memory using the *recorded per-tensor* dtypes.

    Reads are sized per input tensor with its own dtype (so an fp8/int4 weight
    counts 1 byte while a bf16 activation counts 2) — more accurate than a
    single global dtype for quantized ops. The write (output) is sized at the
    activation dtype. Falls back to 0 when shapes aren't concrete.
    """
    if not shapes:
        return 0
    reads = 0
    for i, s in enumerate(shapes):
        n = _prod_ints(s)
        if n == 0:
            return 0
        b = dtype_size(dtypes[i]) if i < len(dtypes) and dtypes[i] else act_bytes
        reads += n * b
    base = op_name.split("::")[-1].lower()
    if (base in _MM_OP_BASES and len(shapes) >= 2
            and len(shapes[0]) >= 2 and len(shapes[1]) >= 2):
        out = _prod_ints(shapes[0][:-1]) * shapes[1][-1]
    else:
        out = _prod_ints(shapes[0])
    return reads + out * act_bytes


def _config_symbols(base_symbols: dict[str, int], cfg: dict) -> dict[str, int]:
    """Return a copy of a profile graph's symbol table with the config
    variables (S/B/C/S+C/TP) overridden for one matrix configuration.

    Config *constants* (H, I, n_h·d, V, ...) are kept as-is so the op template's
    symbolic shapes resolve to this configuration's concrete dims.
    """
    sym = dict(base_symbols)
    if cfg["phase"] == "prefill":
        s_val = int(cfg["seq_len"])
    else:
        s_val = 1  # decode advances each sequence by one token
    c_val = int(cfg.get("ctx_len") or 0)
    sym["S"] = s_val
    sym["B"] = int(cfg["batch_size"])
    sym["C"] = c_val
    sym["S+C"] = s_val + c_val
    sym["TP"] = int(cfg["tp_size"])
    return sym


def _resolve_shape_ints(input_shapes, symbols: dict[str, int]) -> list[list[int]]:
    """Resolve an op's (symbolic) input shapes to concrete integer tensor shapes.

    Only fully-resolvable tensor (list) shapes with all-integer dims are kept;
    scalars and shapes with a dim that can't be resolved to an int are dropped,
    so ``estimate_memory``/``estimate_flops`` see a clean ``list[list[int]]``.
    """
    out: list[list[int]] = []
    for shape in input_shapes or []:
        if not isinstance(shape, list):
            continue
        dims = [_resolve_dim(d, symbols) for d in shape]
        if all(isinstance(d, int) for d in dims):
            out.append(dims)
    return out


def _validate_derived_shapes(template: dict) -> dict:
    """Round-trip check that the symbolic derivation reproduces reality.

    Resolving each op's symbolic ``input_shapes`` at the *profiled* config (the
    template's own ``symbols``) must reproduce the numeric shape recorded in the
    trace (``recorded_shapes``). Context-annotated KV dims (``C`` / ``S+C``) are
    excluded because the context is deliberately added, not recorded; dims that
    don't resolve to an int are skipped (can't compare). Returns counts + a few
    mismatch examples for the Info sheet.
    """
    syms = template.get("symbols", {})
    total = matched = 0
    examples: list[str] = []
    for phase in ("prefill", "decode"):
        tree = template.get(phase)
        if not tree:
            continue
        for node in _flatten_graph_nodes(tree):
            for op in node["ops"]:
                recorded = op.get("recorded_shapes")
                sym = op.get("input_shapes")
                if not recorded or not sym:
                    continue
                for rec, ss in zip(recorded, sym):
                    if not isinstance(ss, list) or not isinstance(rec, list):
                        continue
                    if any(isinstance(d, str) and d in ("C", "S+C") for d in ss):
                        continue  # deliberately context-annotated KV row
                    resolved = [_resolve_dim(d, syms) for d in ss]
                    if not all(isinstance(d, int) for d in resolved):
                        continue  # unresolved dim — nothing to compare
                    total += 1
                    if resolved == list(rec):
                        matched += 1
                    elif len(examples) < 6:
                        examples.append(
                            f"{op.get('name', '')} {ss}\u2192{resolved} vs {rec}")
    return {"total": total, "matched": matched,
            "mismatched": total - matched, "examples": examples}


@app.route("/api/export/shape-matrix", methods=["POST"])
def export_shape_matrix():
    """Export op shapes/dtypes across configurations, grounded in a profiling run.

    Produces a flat Excel table where each row is one
    (Phase, SeqLen, CtxLen, BatchSize, TP, Op) combination.
    Columns for Phase/SeqLen/CtxLen/BatchSize/TP enable Excel filtering
    to select any desired configuration subset.

    Prefill configs: sweep seq_lens × context_lens × batch_sizes × tp_sizes
      (assumes chunked prefill; seq_len = chunk size)
    Decode configs: seq_len fixed to 1, sweep context_lens × batch_sizes × tp_sizes

    The op set + real shapes come from the latest completed profiling run for
    this model (its reconstructed graph is used as a symbolic template): each
    config re-resolves S/B/C/TP and recomputes Memory/FLOPs from the resolved
    shapes (using the recorded per-tensor dtypes). The op set is fixed at the
    profiled config (TP collectives, MoE routing, chunked-prefill splits), so the
    caller profiles at each TP it needs; S/B/C are parametric from one profile.
    The frontend ensures a matching profile exists (reusing the latest run or
    launching a fresh one) before calling this endpoint.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = request.json or {}

    model_id = data.get("model_id")
    if not model_id:
        return jsonify({"ok": False, "error": "No model_id specified"}), 400

    # Prefill settings
    prefill_seq_lens = data.get("prefill_seq_lens",
                                [128, 256, 512, 1024, 2048, 4096, 8192])
    prefill_ctx_lens = data.get("prefill_ctx_lens", [0, 8192])
    prefill_batch_sizes = data.get("prefill_batch_sizes", [1])

    # Decode settings (seq_len always 1)
    decode_ctx_lens = data.get("decode_ctx_lens", [8192])
    decode_batch_sizes = data.get("decode_batch_sizes",
                                  [1, 2, 4, 8, 16, 32, 64, 128])

    # TP sizes
    tp_sizes = data.get("tp_sizes", [1, 2, 4, 8])

    # Validate inputs
    if not isinstance(prefill_seq_lens, list) or not prefill_seq_lens:
        return jsonify({"ok": False,
                        "error": "prefill_seq_lens must be a non-empty list"}), 400
    if not isinstance(tp_sizes, list) or not tp_sizes:
        return jsonify({"ok": False,
                        "error": "tp_sizes must be a non-empty list"}), 400
    if not isinstance(decode_ctx_lens, list) or not decode_ctx_lens:
        return jsonify({"ok": False,
                        "error": "decode_ctx_lens must be a non-empty list"}), 400
    if not isinstance(decode_batch_sizes, list) or not decode_batch_sizes:
        return jsonify({"ok": False,
                        "error": "decode_batch_sizes must be a non-empty list"}), 400

    # Build list of all configurations to sweep (always both phases)
    configs: list[dict] = []
    for seq in prefill_seq_lens:
        for ctx in prefill_ctx_lens:
            for bs in prefill_batch_sizes:
                for tp in tp_sizes:
                    configs.append({
                        "phase": "prefill", "seq_len": seq, "ctx_len": ctx,
                        "batch_size": bs, "tp_size": tp,
                    })
    for ctx in decode_ctx_lens:
        for bs in decode_batch_sizes:
            for tp in tp_sizes:
                configs.append({
                    "phase": "decode", "seq_len": 1, "ctx_len": ctx,
                    "batch_size": bs, "tp_size": tp,
                })

    if not configs:
        return jsonify({"ok": False,
                        "error": "No configurations generated."}), 400

    # Grab the latest reconstructed graph as the op template (real dispatched ops
    # with symbolic shapes) and validate it matches the requested model. The
    # frontend guarantees a matching completed run exists before calling here.
    with _profile_lock:
        state_status = _profile_state["status"]
        state_model = _profile_state.get("model_id")
        state_result = _profile_state.get("result")
        profile_settings = _profile_state.get("settings")
    if state_status != "done" or not state_result:
        return jsonify({
            "ok": False,
            "error": "The Shape Matrix is derived from a profiling run, but no "
                     "completed run is available. Run a profile first.",
        }), 400
    profile_template = state_result.get("graph")
    if not profile_template or not (profile_template.get("prefill")
                                    or profile_template.get("decode")):
        return jsonify({
            "ok": False,
            "error": "The latest profile has no reconstructed graph to "
                     "derive shapes from.",
        }), 400
    if state_model and state_model != model_id:
        return jsonify({
            "ok": False,
            "error": f"Latest profile is for '{state_model}', not '{model_id}'. "
                     "Profile that model or switch the model ID.",
        }), 400

    pdtype_bytes = profile_template.get("config", {}).get("dtype_bytes", 2)

    # Estimate row count (configs × ~ops_per_config) for the limit guard.
    test_tree = profile_template.get("prefill") or profile_template.get("decode")
    test_ops_count = 0
    if test_tree:
        for node in _flatten_graph_nodes(test_tree):
            test_ops_count += len(node["ops"])
    estimated_rows = len(configs) * test_ops_count
    if estimated_rows > _MAX_MATRIX_ROWS:
        return jsonify({
            "ok": False,
            "error": f"Too many rows ({estimated_rows}). Max is {_MAX_MATRIX_ROWS}. "
                     "Reduce seq_lens, batch_sizes, ctx_lens, or tp_sizes."
        }), 400

    # Build flat table data
    wb = Workbook()

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E",
                              fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

    # Sheet name: use model short name (last part of model_id)
    model_short = model_id.split("/")[-1] if "/" in model_id else model_id
    # Sanitize for Excel sheet name (max 31 chars, no special chars)
    sheet_name = model_short[:31].replace("[", "").replace("]", "")

    ws = wb.active
    ws.title = sheet_name

    headers = [
        "Phase", "Seq Len", "Ctx Len", "Batch Size", "TP",
        "Module", "Op Name", "Backend", "Layers",
        "Symbolic Shape", "Shape",
        "Memory (bytes)", "FLOPs", "AI",
    ]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(1, col, hdr)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    graph_cfg = profile_template.get("config", {})
    base_symbols = profile_template.get("symbols", {})
    for cfg in configs:
        tree = profile_template.get(cfg["phase"])
        if not tree:
            continue
        symbols = _config_symbols(base_symbols, cfg)

        flat_nodes = _flatten_graph_nodes(tree)
        for node_info in flat_nodes:
            effective_repeat = node_info.get("effective_repeat", 1)
            for op in node_info["ops"]:
                # Prefer the real per-tensor dtypes recorded in the trace.
                recorded_dtypes = op.get("input_dtypes")
                shape_str = _format_op_shape_with_dtypes(
                    op, symbols, graph_cfg, recorded_dtypes=recorded_dtypes)

                # Symbolic shape: keep only config variables (S, B, C, TP)
                # symbolic, resolve model constants to numbers.
                sym_shapes = op.get("input_shapes", [])
                if sym_shapes:
                    sym_parts = []
                    for s in sym_shapes:
                        if isinstance(s, list):
                            dims = [_partially_resolve_dim(d, symbols)
                                    for d in s]
                            sym_parts.append("[" + ", ".join(dims) + "]")
                        else:
                            sym_parts.append(
                                _partially_resolve_dim(s, symbols))
                    symbolic_str = " × ".join(sym_parts)
                else:
                    symbolic_str = "—"

                # Recompute Memory/FLOPs from this config's resolved shapes so
                # every row is self-consistent; Memory uses the recorded
                # per-tensor dtypes for accuracy.
                resolved = _resolve_shape_ints(sym_shapes, symbols)
                op_name = op.get("name", "")
                if recorded_dtypes:
                    mem_bytes = _profile_op_memory(
                        op_name, resolved, recorded_dtypes, pdtype_bytes)
                else:
                    mem_bytes = estimate_memory(op_name, resolved, pdtype_bytes)
                flops = estimate_flops(op_name, resolved)
                ai = round(flops / mem_bytes, 2) if mem_bytes > 0 else 0

                # Merge module path and op role into single column
                path = node_info["path"]
                role = op.get("role", "")
                module_col = f"{path}.{role}" if role else path

                ws.cell(row, 1, cfg["phase"])
                ws.cell(row, 2, cfg["seq_len"])
                ws.cell(row, 3, cfg["ctx_len"])
                ws.cell(row, 4, cfg["batch_size"])
                ws.cell(row, 5, cfg["tp_size"])
                ws.cell(row, 6, module_col)
                ws.cell(row, 7, op.get("name", ""))
                ws.cell(row, 8, op.get("backend", ""))
                ws.cell(row, 9, effective_repeat)
                ws.cell(row, 10, symbolic_str)
                ws.cell(row, 11, shape_str)
                ws.cell(row, 12, mem_bytes)
                ws.cell(row, 13, flops)
                ws.cell(row, 14, ai)

                for c in range(1, len(headers) + 1):
                    ws.cell(row, c).border = thin_border
                row += 1

    # AutoFit column widths by sampling header + first/last 100 data rows
    sample_rows = list(range(1, min(row, 102)))  # header + first 100
    if row > 202:
        sample_rows += list(range(row - 100, row))  # last 100
    elif row > 102:
        sample_rows += list(range(102, row))
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for r in sample_rows:
            val = ws.cell(r, col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    # Provenance / caveats sheet so the optimization engineers know exactly
    # which profiling run the shapes were grounded in.
    info = wb.create_sheet("Info")
    ps = profile_settings or {}
    pcfg = profile_template.get("config", {})
    info_rows = [
        ("Shape source", "profile-derived (grounded in a profiling run)"),
        ("Model", model_id),
        ("Profiled query_len (S)", ps.get("query_len")),
        ("Profiled context_len (C)", ps.get("context_len")),
        ("Profiled decode batch (B)", ps.get("decode_batch_size")),
        ("Profiled TP", pcfg.get("tp_size", ps.get("tp_size"))),
        ("Profiled quantization", pcfg.get("quantization")),
        ("Profiled mode", ps.get("mode")),
        ("", ""),
        ("How rows are derived",
         "The profile contributes the accurate op set, real recorded "
         "shapes and backends. Shapes are re-resolved per config "
         "(S/B/C/TP). Memory/FLOPs are then analytic functions of "
         "(op, shape, dtype), recomputed per config — NOT measured values."),
        ("Memory/FLOPs are estimates",
         "Heuristic (op+shape) estimates, not measured; e.g. attention "
         "FLOPs are not modeled. Only the op set and shapes come from the "
         "trace. Measured device time is intentionally not in this sweep."),
        ("Caveat — op set / TP",
         "The op set (TP collectives, MoE routing, chunked-prefill splits) "
         "is fixed at the profiled config. Profile at each TP you need — "
         "sweeping TP only divides /TP dims, it does not add comm ops that "
         "weren't profiled."),
        ("Context (C) is parametric",
         "No need to profile per context. One base profile with any "
         "non-zero context captures C as S+C on the prefill attention KV "
         "rows; every other context length is then derived by resolving C. "
         "A base profile with context=0 leaves KV rows as S, so context "
         "can't be derived — profile with a small non-zero context."),
        ("Caveat — timing",
         "Device time is valid only at the profiled point and is not "
         "included in this sweep; rows carry shapes/memory/FLOPs only."),
    ]
    # Round-trip validation: re-resolving each op's symbolic shape at the
    # profiled config must reproduce the shape actually recorded in the
    # trace — a self-consistency proof for the derivation machinery.
    vres = _validate_derived_shapes(profile_template)
    if vres["total"]:
        pct = 100.0 * vres["matched"] / vres["total"]
        info_rows.append((
            "Shape validation",
            f"{vres['matched']}/{vres['total']} op input shapes "
            f"({pct:.1f}%) re-resolve at the profiled config to exactly the "
            "shape recorded in the trace (context-annotated KV rows "
            "excluded). This validates the symbolic derivation.",
        ))
        if vres["mismatched"]:
            info_rows.append((
                "Validation mismatches",
                f"{vres['mismatched']} mismatched. Examples: "
                + " | ".join(vres["examples"]),
            ))
    for r_idx, (k, v) in enumerate(info_rows, 1):
        kc = info.cell(r_idx, 1, k)
        kc.font = Font(bold=True, size=10)
        kc.alignment = Alignment(vertical="top")
        vc = info.cell(r_idx, 2, "" if v is None else str(v))
        vc.alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 90

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    model_name = model_id.replace("/", "_")
    quant_tag = pcfg.get("quantization") or "none"
    filename = f"vllm_xpu_shape_matrix_{model_name}_{quant_tag}_profile.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="vLLM Breakdown Web UI")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8080)
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    print(f"Starting vLLM Breakdown at http://{args.host}:{args.port}")
    print(f"Open http://localhost:{args.port} in your browser")
    app.run(host=args.host, port=args.port, debug=args.debug)
