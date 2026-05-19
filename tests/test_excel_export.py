#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Unit tests for the Excel export endpoint.

Tests cover:
  - Summary sheet with numeric config values for formula references
  - Operations sheet with symbolic and concrete shape columns
  - Formula generation referencing Summary sheet cells
  - Model Hierarchy sheet reflecting graph tree structure
  - Backward compatibility when no graph data is provided
"""

from __future__ import annotations

import io
import json
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app
from breakdown.model_graph import build_model_graph


MOCK_SUMMARY = {
    "architecture": "Qwen3ForCausalLM",
    "model_type": "qwen3",
    "hidden_size": 2560,
    "num_layers": 36,
    "num_heads": 32,
    "num_kv_heads": 8,
    "head_dim": 80,
    "intermediate_size": 6912,
    "vocab_size": 151936,
    "dtype": "bfloat16",
    "is_moe": False,
    "num_experts": None,
    "num_experts_per_tok": None,
    "quant_method": None,
}


def _build_test_payload(include_graph: bool = True) -> dict:
    """Build a test payload for the Excel export endpoint."""
    ops = [
        {
            "name": "aten::mm",
            "backend": "torch-xpu-ops",
            "input_shapes": [["S", "H"], ["H", "n_h\u00b7d"]],
            "dtype": "bfloat16",
            "call_count": 32,
            "layer_count": 36,
            "device_time_us": 32000,
            "cpu_time_us": 1200,
            "memory_bytes": 2621440,
            "flops": 335544320,
        },
        {
            "name": "rms_norm",
            "backend": "vllm-xpu-kernels",
            "input_shapes": [["S", "H"], ["H"]],
            "dtype": "bfloat16",
            "call_count": 64,
            "layer_count": 36,
            "device_time_us": 8000,
            "cpu_time_us": 500,
            "memory_bytes": 655360,
            "flops": 655360,
        },
        {
            "name": "triton_flash_attn",
            "backend": "triton",
            "input_shapes": [["S", "n_h", "d"]],
            "dtype": "bfloat16",
            "call_count": 32,
            "layer_count": 36,
            "device_time_us": 20000,
            "cpu_time_us": 1500,
            "memory_bytes": 1310720,
            "flops": 167772160,
        },
    ]

    data = {
        "model_id": "Qwen/Qwen3-4B",
        "mode": "eager",
        "summary": MOCK_SUMMARY,
        "total_device_time_us": 60000,
        "total_cpu_time_us": 3200,
        "backends": {
            "torch-xpu-ops": {
                "device_time_us": 32000,
                "pct": 53.3,
                "num_ops": 1,
                "num_calls": 32,
            },
            "vllm-xpu-kernels": {
                "device_time_us": 8000,
                "pct": 13.3,
                "num_ops": 1,
                "num_calls": 64,
            },
            "triton": {
                "device_time_us": 20000,
                "pct": 33.3,
                "num_ops": 1,
                "num_calls": 32,
            },
        },
        "ops": ops,
    }

    if include_graph:
        graph = build_model_graph(
            MOCK_SUMMARY, prefill_len=128, decode_batch=1, context_len=4096
        )
        data["graph"] = graph
        data["phase"] = "prefill"

    return data


class TestExcelExport(unittest.TestCase):
    """Tests for /api/export/excel endpoint."""

    def setUp(self):
        self.client = app.test_client()

    def _export(self, data: dict):
        """Helper to call the export endpoint and return the workbook."""
        from openpyxl import load_workbook

        resp = self.client.post(
            "/api/export/excel",
            data=json.dumps(data),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.data))
        return wb

    def test_sheets_with_graph(self):
        """When graph data is provided, all three sheets should exist."""
        wb = self._export(_build_test_payload(include_graph=True))
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Operations", wb.sheetnames)
        self.assertIn("Model Hierarchy", wb.sheetnames)

    def test_sheets_without_graph(self):
        """Without graph data, Model Hierarchy sheet should not exist."""
        wb = self._export(_build_test_payload(include_graph=False))
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Operations", wb.sheetnames)
        self.assertNotIn("Model Hierarchy", wb.sheetnames)

    def test_summary_config_numeric_values(self):
        """Config values in Summary sheet should be numeric for formula use."""
        wb = self._export(_build_test_payload())
        ws = wb["Summary"]

        # Find hidden_size and verify it's stored as a number
        for r in range(5, 25):
            if ws.cell(r, 1).value == "hidden_size":
                val = ws.cell(r, 2).value
                self.assertEqual(val, 2560)
                self.assertIsInstance(val, int)
                return
        self.fail("hidden_size not found in Summary sheet")

    def test_operations_shape_columns(self):
        """Operations sheet should have both symbolic and concrete shape cols."""
        wb = self._export(_build_test_payload())
        ws = wb["Operations"]

        headers = [ws.cell(1, c).value for c in range(1, 13)]
        self.assertIn("Shape (symbolic)", headers)
        self.assertIn("Shape (concrete)", headers)

    def test_concrete_shape_formula_references_summary(self):
        """Concrete shapes with symbolic dims should use Summary references."""
        data = _build_test_payload(include_graph=False)
        # Use single-shape op to get formula
        data["ops"] = [
            {
                "name": "test_single_shape",
                "backend": "torch-xpu-ops",
                "input_shapes": [["S", "H"]],
                "dtype": "bfloat16",
                "call_count": 1,
                "layer_count": 1,
                "device_time_us": 10,
                "memory_bytes": 100,
                "flops": 200,
            }
        ]
        wb = self._export(data)
        ws = wb["Operations"]

        # Column 4 = Shape (concrete)
        cell_value = ws.cell(2, 4).value
        self.assertTrue(
            cell_value.startswith("="),
            f"Expected formula, got: {cell_value}",
        )
        self.assertIn("Summary!", cell_value)

    def test_concrete_shape_no_formula_for_numeric(self):
        """All-numeric shapes should not generate formulas."""
        data = _build_test_payload(include_graph=False)
        data["ops"] = [
            {
                "name": "test_numeric",
                "backend": "torch-xpu-ops",
                "input_shapes": [[128, 2560], [2560, 2560]],
                "dtype": "bfloat16",
                "call_count": 1,
                "layer_count": 1,
                "device_time_us": 10,
                "memory_bytes": 100,
                "flops": 200,
            }
        ]
        wb = self._export(data)
        ws = wb["Operations"]

        cell_value = ws.cell(2, 4).value
        self.assertFalse(
            str(cell_value).startswith("="),
            f"All-numeric should not be formula, got: {cell_value}",
        )
        self.assertIn("128", str(cell_value))
        self.assertIn("2560", str(cell_value))

    def test_hierarchy_reflects_model_structure(self):
        """Hierarchy sheet should show nested modules and their ops."""
        wb = self._export(_build_test_payload())
        ws = wb["Model Hierarchy"]

        # First data row should be the root "model" module
        self.assertEqual(ws.cell(2, 1).value, "model")

        # Should find op rows (col 8 has op role)
        found_op = False
        for r in range(2, 100):
            if ws.cell(r, 8).value:
                found_op = True
                # Op should have shape info (col 11)
                self.assertIsNotNone(ws.cell(r, 11).value)
                break
        self.assertTrue(found_op, "No ops found in hierarchy sheet")

    def test_hierarchy_concrete_shapes(self):
        """Op shapes in hierarchy should show concrete numeric values."""
        wb = self._export(_build_test_payload())
        ws = wb["Model Hierarchy"]

        # Find first op with a shape (col 11) and verify it has correct numbers
        for r in range(2, 100):
            shape = ws.cell(r, 11).value
            if shape and shape != "—":
                # Should contain numbers (from resolved symbols)
                self.assertRegex(
                    shape, r"\d+",
                    f"Shape should contain numbers: {shape}",
                )
                # Verify that H=2560 appears (hidden_size from MOCK_SUMMARY)
                if "2560" in shape:
                    return
        # At minimum some shape should have been found with 2560
        # (embedding op uses [V, H] = [151936, 2560])
        self.fail("No concrete shape with hidden_size=2560 found")

    def test_export_error_no_ops(self):
        """Should return 400 when no ops data provided."""
        resp = self.client.post(
            "/api/export/excel",
            data=json.dumps({"model_id": "test"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 400)


if __name__ == "__main__":
    unittest.main()
