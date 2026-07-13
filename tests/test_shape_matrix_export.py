#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Unit tests for the Shape Matrix Export endpoint.

Tests cover:
  - Flat table structure with config columns for filtering
  - Separate prefill/decode settings (always both phases)
  - TP size sweeping affects shapes
  - Decode seq_len fixed at 1
  - Dtype included in shapes
  - Symbolic shapes, Memory, FLOPs, AI columns
  - Sheet named after model
  - Quantization in filename
  - Validation: no model_id, too many rows, empty arrays
"""

from __future__ import annotations

import io
import json
import os
import sys
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app


MOCK_QWEN3_CONFIG = {
    "architectures": ["Qwen3ForCausalLM"],
    "model_type": "qwen3",
    "hidden_size": 2560,
    "num_hidden_layers": 36,
    "num_attention_heads": 32,
    "num_key_value_heads": 8,
    "head_dim": 80,
    "intermediate_size": 6912,
    "vocab_size": 151936,
    "torch_dtype": "bfloat16",
}


class TestShapeMatrixExport(unittest.TestCase):
    """Tests for /api/export/shape-matrix endpoint (always profile-derived)."""

    def setUp(self):
        self.client = app.test_client()
        # The Shape Matrix is derived from a completed profiling run; inject one.
        import app as app_module
        self._saved_state = app_module._profile_state
        app_module._profile_state = {
            "status": "done",
            "result": {"graph": _mock_profile_graph()},
            "error": None,
            "model_id": "Qwen/Qwen3-4B",
            "settings": {"query_len": 128, "context_len": 0,
                         "decode_batch_size": 8, "tp_size": 1, "mode": "eager"},
        }

    def tearDown(self):
        import app as app_module
        app_module._profile_state = self._saved_state

    def _export(self, data: dict):
        """Helper to call the endpoint."""
        return self.client.post(
            "/api/export/shape-matrix",
            data=json.dumps(data),
            content_type="application/json",
        )

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_basic_export(self, mock_fetch):
        """Basic export produces a flat table with correct headers."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [4096],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        headers = [ws.cell(1, c).value for c in range(1, 15)]
        self.assertEqual(headers, [
            "Phase", "Seq Len", "Ctx Len", "Batch Size", "TP",
            "Module", "Op Name", "Backend", "Layers",
            "Symbolic Shape", "Shape",
            "Memory (bytes)", "FLOPs", "AI",
        ])

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_sheet_named_after_model(self, mock_fetch):
        """Sheet should be named after the model."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        self.assertEqual(wb.sheetnames[0], "Qwen3-4B")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_quantization_in_filename(self, mock_fetch):
        """Filename should include quantization tag."""
        import app as app_module
        # The Shape Matrix reuses the last run only when the quantization
        # matches, so profile settings must report the same fp8 quantization.
        app_module._profile_state["settings"]["quantization"] = "fp8"
        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
            "quantization": "fp8",
        })
        cd = resp.headers.get("Content-Disposition", "")
        self.assertIn("fp8", cd)

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_quantization_mismatch_rejected(self, mock_fetch):
        """Requesting a quantization the profile didn't use is rejected."""
        # Default injected run has no quantization; requesting fp8 must 400.
        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
            "quantization": "fp8",
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("quantization", resp.get_json()["error"].lower())

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_always_both_phases(self, mock_fetch):
        """Both prefill and decode are always included."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        phases = set()
        for r in range(2, 500):
            p = ws.cell(r, 1).value
            if p is None:
                break
            phases.add(p)
        self.assertEqual(phases, {"prefill", "decode"})

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_decode_seq_len_fixed_at_1(self, mock_fetch):
        """Decode phase always has seq_len=1."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [2048],
            "prefill_ctx_lens": [4096],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1, 8],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        for r in range(2, 500):
            if ws.cell(r, 1).value is None:
                break
            if ws.cell(r, 1).value == "decode":
                self.assertEqual(ws.cell(r, 2).value, 1)

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_tp_affects_shapes(self, mock_fetch):
        """Different TP sizes produce different shapes."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [4096],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1, 4],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]

        tp1_shape = tp4_shape = None
        for r in range(2, 500):
            if ws.cell(r, 1).value is None:
                break
            if ws.cell(r, 6).value and "qkv_proj" in ws.cell(r, 6).value and ws.cell(r, 1).value == "prefill":
                if ws.cell(r, 5).value == 1:
                    tp1_shape = ws.cell(r, 11).value
                elif ws.cell(r, 5).value == 4:
                    tp4_shape = ws.cell(r, 11).value

        self.assertIsNotNone(tp1_shape)
        self.assertIsNotNone(tp4_shape)
        self.assertNotEqual(tp1_shape, tp4_shape)

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_symbolic_shapes_column(self, mock_fetch):
        """Symbolic Shape column shows dimension names."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]

        # Find qkv_proj - its symbolic shape should have "S" symbolic and
        # model constants resolved to numbers
        for r in range(2, 200):
            if ws.cell(r, 6).value and "qkv_proj" in ws.cell(r, 6).value:
                sym = ws.cell(r, 10).value
                # S stays symbolic, H is resolved to number
                self.assertIn("S", sym)
                self.assertIn("\u00d7", sym)
                # H (2560) should be resolved to its numeric value
                self.assertIn("2560", sym)
                return
        self.fail("qkv_proj not found")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_memory_flops_ai_columns(self, mock_fetch):
        """Memory, FLOPs, and AI columns have values for matmul ops."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]

        for r in range(2, 200):
            if ws.cell(r, 6).value and "qkv_proj" in ws.cell(r, 6).value:
                mem = ws.cell(r, 12).value
                flops = ws.cell(r, 13).value
                ai = ws.cell(r, 14).value
                self.assertGreater(mem, 0)
                self.assertGreater(flops, 0)
                self.assertGreater(ai, 0)
                return
        self.fail("qkv_proj not found")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_dtype_in_shapes(self, mock_fetch):
        """Shapes include dtype information (bf16, fp8, etc.)."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]

        found_dtype = False
        for r in range(2, 100):
            val = ws.cell(r, 11).value
            if val and "bf16" in str(val):
                found_dtype = True
                break
        self.assertTrue(found_dtype, "Expected bf16 dtype in shapes")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_freeze_panes_and_filter(self, mock_fetch):
        """Sheet should have frozen panes and auto-filter."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128],
            "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1],
            "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1],
            "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_error_no_model_id(self):
        """Should return 400 when no model_id specified."""
        resp = self._export({})
        self.assertEqual(resp.status_code, 400)
        self.assertIn("model_id", resp.json["error"])

    def test_error_empty_tp_sizes(self):
        """Should return 400 when tp_sizes is empty."""
        resp = self._export({
            "model_id": "test",
            "tp_sizes": [],
        })
        self.assertEqual(resp.status_code, 400)

    def test_error_empty_decode_ctx_lens(self):
        """Should return 400 when decode_ctx_lens is empty."""
        resp = self._export({
            "model_id": "test",
            "decode_ctx_lens": [],
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("decode_ctx_lens", resp.json["error"])

    def test_error_empty_decode_batch_sizes(self):
        """Should return 400 when decode_batch_sizes is empty."""
        resp = self._export({
            "model_id": "test",
            "decode_batch_sizes": [],
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("decode_batch_sizes", resp.json["error"])

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_error_too_many_rows(self, mock_fetch):
        """Should return 400 when sweep produces too many rows."""
        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": list(range(1, 100)),
            "prefill_ctx_lens": list(range(1, 100)),
            "prefill_batch_sizes": list(range(1, 20)),
            "tp_sizes": [1, 2, 4, 8],
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("Too many rows", resp.json["error"])

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_default_values(self, mock_fetch):
        """Endpoint works with minimal input (defaults applied)."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        self.assertIsNotNone(ws.cell(2, 1).value)


def _mock_profile_graph():
    """A minimal profile-reconstructed graph template (matches the serialized
    shape of ``build_graph_from_trace``) for the profile-derived export path."""
    def op(name, role, shapes, recorded, dtypes):
        return {"name": name, "role": role, "backend": "torch-xpu-ops",
                "input_shapes": shapes, "recorded_shapes": recorded,
                "input_dtypes": dtypes, "memory_bytes": 111, "flops": 222}

    def leaf(name, cls, path, ops):
        return {"name": name, "module_type": cls, "path": path,
                "repeat_count": 1, "ops": ops, "children": []}

    prefill = {
        "name": "model", "module_type": "Qwen3Model", "path": "model",
        "repeat_count": 1, "ops": [], "children": [
            leaf("self_attn", "Attention", "model/attn", [
                op("aten::mm", "qkv_proj", [["S", "H"], ["H", "QKV/TP"]],
                   [[128, 2560], [2560, 3840]],
                   ["bfloat16", "float8_e4m3fn"]),
                op("vllm::unified_attention_with_output", "attention",
                   [["S", "n_h/TP", "d"], ["S+C", "n_kv/TP", "d"],
                    ["S+C", "n_kv/TP", "d"]],
                   [[128, 32, 80], [128, 8, 80], [128, 8, 80]],
                   ["bfloat16", "bfloat16", "bfloat16"]),
            ]),
        ],
    }
    decode = {
        "name": "model", "module_type": "Qwen3Model", "path": "model",
        "repeat_count": 1, "ops": [], "children": [
            leaf("self_attn", "Attention", "model/attn", [
                op("aten::mm", "qkv_proj", [["B", "H"], ["H", "QKV/TP"]],
                   [[1, 2560], [2560, 3840]],
                   ["bfloat16", "float8_e4m3fn"]),
            ]),
        ],
    }
    return {
        "prefill": prefill, "decode": decode,
        "symbols": {"H": 2560, "n_h": 32, "n_kv": 8, "d": 80, "I": 6912,
                    "V": 151936, "n_h\u00b7d": 2560, "QKV": 3840, "TP": 1,
                    "S": 128, "B": 1, "C": 0, "S+C": 128},
        "config": {"tp_size": 1, "quantization": "fp8", "dtype_bytes": 2,
                   "weight_dtype_bytes": 1, "num_layers": 36},
        "source": "profile",
    }


class TestShapeMatrixProfileSource(unittest.TestCase):
    """Tests for the profile-derived Shape Matrix export (source='profile')."""

    def setUp(self):
        self.client = app.test_client()
        # Inject a completed profiling run as the derivation template.
        import app as app_module
        self._saved_state = app_module._profile_state
        app_module._profile_state = {
            "status": "done",
            "result": {"graph": _mock_profile_graph()},
            "error": None,
            "model_id": "Qwen/Qwen3-4B",
            "settings": {"query_len": 128, "context_len": 0,
                         "decode_batch_size": 8, "tp_size": 1, "mode": "eager"},
        }

    def tearDown(self):
        import app as app_module
        app_module._profile_state = self._saved_state

    def _export(self, data: dict):
        return self.client.post(
            "/api/export/shape-matrix",
            data=json.dumps(data),
            content_type="application/json",
        )

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_derived_basic(self, mock_fetch):
        """Profile source produces rows and an Info provenance sheet."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.data))
        self.assertIn("Info", wb.sheetnames)
        ws = wb["Qwen3-4B"]
        # The dispatched attention op from the trace should be present (it is
        # not something the config-driven builder would name identically).
        names = {ws.cell(r, 7).value for r in range(2, 200)
                 if ws.cell(r, 7).value}
        self.assertIn("vllm::unified_attention_with_output", names)
        self.assertIn("fp8", resp.headers.get("Content-Disposition", ""))

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_tp_divides_shapes(self, mock_fetch):
        """Sweeping TP re-resolves the /TP dims of the profiled template."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1, 2],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        tp1 = tp2 = None
        for r in range(2, 200):
            if ws.cell(r, 1).value is None:
                break
            if (ws.cell(r, 1).value == "prefill"
                    and "qkv_proj" in (ws.cell(r, 6).value or "")):
                if ws.cell(r, 5).value == 1:
                    tp1 = ws.cell(r, 11).value
                elif ws.cell(r, 5).value == 2:
                    tp2 = ws.cell(r, 11).value
        self.assertIsNotNone(tp1)
        self.assertIsNotNone(tp2)
        # QKV=3840 at TP=1, 1920 at TP=2.
        self.assertIn("3840", tp1)
        self.assertIn("1920", tp2)

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_context_rewrites_kv(self, mock_fetch):
        """A prefill context sweep resolves attention KV rows to S+C."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [4096],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        found = False
        for r in range(2, 200):
            if (ws.cell(r, 7).value == "vllm::unified_attention_with_output"
                    and ws.cell(r, 1).value == "prefill"):
                # KV rows must read S+C = 128+4096 = 4224, query stays 128.
                self.assertIn("4224", ws.cell(r, 11).value)
                self.assertIn("128", ws.cell(r, 11).value)
                found = True
                break
        self.assertTrue(found, "attention op not found")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_memory_recomputed_per_config(self, mock_fetch):
        """Memory differs across configs (recomputed, not the template's)."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128, 512], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        mems = {}
        for r in range(2, 200):
            if ws.cell(r, 1).value is None:
                break
            if (ws.cell(r, 1).value == "prefill"
                    and "qkv_proj" in (ws.cell(r, 6).value or "")):
                mems[ws.cell(r, 2).value] = ws.cell(r, 12).value
        self.assertIn(128, mems)
        self.assertIn(512, mems)
        # Neither equals the template placeholder (111) and they differ by S.
        self.assertNotEqual(mems[128], 111)
        self.assertNotEqual(mems[128], mems[512])

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_recorded_dtypes(self, mock_fetch):
        """The Shape column uses the real per-tensor dtypes from the trace."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        found = False
        for r in range(2, 200):
            if (ws.cell(r, 1).value == "prefill"
                    and "qkv_proj" in (ws.cell(r, 6).value or "")):
                shape = ws.cell(r, 11).value
                # activation is bf16, the weight tensor is the recorded fp8.
                self.assertIn("bf16", shape)
                self.assertIn("fp8", shape)
                found = True
                break
        self.assertTrue(found, "qkv_proj not found")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_memory_uses_per_tensor_dtype(self, mock_fetch):
        """fp8 weight is sized at 1 byte, not the activation's 2 bytes."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb["Qwen3-4B"]
        for r in range(2, 200):
            if (ws.cell(r, 1).value == "prefill"
                    and "qkv_proj" in (ws.cell(r, 6).value or "")):
                # act 128*2560*2 + weight 2560*3840*1 + out 128*3840*2
                expected = 128 * 2560 * 2 + 2560 * 3840 * 1 + 128 * 3840 * 2
                self.assertEqual(ws.cell(r, 12).value, expected)
                return
        self.fail("qkv_proj not found")

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_shape_validation_info(self, mock_fetch):
        """The Info sheet reports a shape round-trip validation summary."""
        from openpyxl import load_workbook

        resp = self._export({
            "model_id": "Qwen/Qwen3-4B",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        wb = load_workbook(io.BytesIO(resp.data))
        info = wb["Info"]
        summary = None
        for r in info.iter_rows(values_only=True):
            if r[0] and "validation" in str(r[0]).lower():
                summary = str(r[1])
                break
        self.assertIsNotNone(summary, "no validation row in Info sheet")
        # All non-context dims must round-trip at the profiled config.
        self.assertIn("100.0%", summary)

    def test_profile_requires_completed_run(self):
        """source=profile with no completed run returns 400."""
        import app as app_module
        app_module._profile_state = {
            "status": "idle", "result": None, "error": None,
            "model_id": None, "settings": None,
        }
        with patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG):
            resp = self._export({
                "model_id": "Qwen/Qwen3-4B",
                "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
                "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
                "decode_batch_sizes": [1], "tp_sizes": [1],
            })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("profil", resp.json["error"].lower())

    @patch("app.fetch_model_config", return_value=MOCK_QWEN3_CONFIG)
    def test_profile_model_mismatch(self, mock_fetch):
        """source=profile errors when the profiled model differs."""
        resp = self._export({
            "model_id": "other/model",
            "prefill_seq_lens": [128], "prefill_ctx_lens": [0],
            "prefill_batch_sizes": [1], "decode_ctx_lens": [4096],
            "decode_batch_sizes": [1], "tp_sizes": [1],
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("other/model", resp.json["error"])


if __name__ == "__main__":
    unittest.main()

