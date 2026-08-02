# SPDX-License-Identifier: Apache-2.0
"""Ranking, timing statistics, budgets and history - all GPU-free."""
from __future__ import annotations

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from breakdown.bench import (  # noqa: E402
    estimate, history, rank, reports, spec, timing,
)


def _rec(op, latency, layers=1, flops=0.0, nbytes=0.0, phase="decode",
         backend="vllm-xpu-kernels", traced=0.0, comparable=False, **kw):
    r = {
        "op": op, "status": "ok", "device": "xpu", "phase": phase,
        "seq_len": 1, "ctx_len": 2048, "batch_size": 32, "tp": 4,
        "backend": backend, "layers": layers, "latency_us": latency,
        "flops": flops, "bytes": nbytes, "shape": "[32, 6144]",
        "shape_key": f"{op}-key", "case_id": f"{op}-case",
        "traced_device_time_us": traced, "traced_comparable": comparable,
    }
    r.update(kw)
    return r


#: A DRAM-only roofline: no cache roof, so a case is always charged to DRAM.
PEAKS = {"bw_gbs": 456.0, "tflops": 98.3}

#: The real BMG roofline, cache roof included.
CACHE_PEAKS = {"bw_gbs": 456.0, "tflops": 98.3,
               "cache_bytes": 18 * 1024 ** 2, "cache_bw_gbs": 1200.0}


class TestUtilization(unittest.TestCase):
    def test_memory_bound_utilization_is_against_the_bandwidth_peak(self):
        # 456 GB/s for 1 us moves 456e3 bytes; half of that is 50 % of peak.
        util, bound = estimate.utilization(1.0, 0.0, 228_000, PEAKS)
        self.assertEqual(bound, "memory")
        self.assertAlmostEqual(util, 0.5, places=2)

    def test_the_bound_comes_from_arithmetic_intensity_not_the_measurement(self):
        # Machine balance is 98.3e12 / 456e9 = 215 flop/byte. An op above it is
        # compute-bound *however well or badly the kernel ran* - the old
        # "whichever utilization is larger" rule labelled a GEMM running at 30 %
        # of peak FLOPS "memory-bound" and a pure gather "compute-bound".
        ridge = estimate.ridge_ai(PEAKS)
        self.assertAlmostEqual(ridge, 98.3e12 / 456e9, places=3)

        # A GEMM: AI well above the ridge, but only a third of peak FLOPS.
        flops, nbytes = 300.0 * 1e6, 1e6
        util, bound = estimate.utilization(10.0, flops, nbytes, PEAKS)
        self.assertEqual(bound, "compute")
        self.assertAlmostEqual(util, (flops / 10e-6) / 98.3e12, places=3)

        # A gather: two bytes of traffic per flop, far below the ridge.
        util, bound = estimate.utilization(1.0, 1_000.0, 456_000, PEAKS)
        self.assertEqual(bound, "memory")
        self.assertAlmostEqual(util, 1.0, places=2)

    def test_a_cache_resident_op_is_measured_against_the_cache_roof(self):
        # 1 MB fits the 18 MB LLC; the benchmark repeats the kernel on the same
        # operands inside one window, so it is served by the cache. Charging it
        # to DRAM produced "300 % of peak" nonsense.
        nbytes = 1024 ** 2
        util, bound, level = estimate.utilization_detail(
            1.0, 0.0, nbytes, CACHE_PEAKS)
        self.assertEqual((bound, level), ("memory", "cache"))
        self.assertAlmostEqual(util, (nbytes / 1e-6) / 1200e9, places=3)
        # ... and would have been reported as >2x of the DRAM peak.
        self.assertGreater(estimate.utilization(1.0, 0.0, nbytes, PEAKS)[0], 2.0)

    def test_an_op_larger_than_the_cache_is_charged_to_dram(self):
        nbytes = 64 * 1024 ** 2
        _, bound, level = estimate.utilization_detail(1.0, 0.0, nbytes,
                                                      CACHE_PEAKS)
        self.assertEqual((bound, level), ("memory", "dram"))


class TestRank(unittest.TestCase):
    def test_call_count_outweighs_a_single_expensive_op(self):
        recs = [_rec("small_op_in_every_layer", 10.0, layers=57,
                     nbytes=1_000),
                _rec("big_op_once", 400.0, layers=1, nbytes=1_000)]
        doc = rank.rank(recs)
        self.assertEqual(doc["targets"][0]["op"], "small_op_in_every_layer")
        self.assertEqual(doc["schema_version"], rank.SCHEMA_VERSION)
        self.assertEqual(doc["engine"], "replay")

    def test_an_op_at_the_roofline_is_not_a_target(self):
        # 1 us at the 1200 GB/s cache roof = 1.2e6 bytes; 1.1e6 is ~90 % of it.
        recs = [_rec("saturated", 1.0, layers=1, nbytes=1_100_000)]
        doc = rank.rank(recs)
        t = doc["targets"][0]
        self.assertEqual(t["action"], "at_roofline")
        self.assertEqual(t["savings_us"]["total"], 0.0)

    def test_op_with_no_editable_source_is_tune_config_not_a_kernel_session(self):
        recs = [_rec("aten::div", 10.0, backend="torch-xpu-ops", nbytes=1_000)]
        doc = rank.rank(recs)
        self.assertEqual(doc["targets"][0]["action"], "tune_config")

    def test_op_with_editable_source_and_headroom_is_a_kernel_session(self):
        recs = [_rec("_C::silu_and_mul_with_clamp", 100.0, layers=57,
                     nbytes=1_000)]
        doc = rank.rank(recs, kernel_sources=rank.load_kernel_sources())
        t = doc["targets"][0]
        self.assertEqual(t["action"], "optimize_kernel")
        self.assertTrue(t["kernel"]["build_cmd"])
        self.assertGreater(t["savings_us"]["total"], 0)

    def test_impossible_utilization_flags_the_cost_model_not_the_kernel(self):
        # The analytic bytes charge an embedding for the whole table it could
        # read; the resulting 300x-of-peak "utilization" must not retire the op
        # as done.
        recs = [_rec("aten::embedding", 1.0, nbytes=456_000_00 * 300)]
        doc = rank.rank(recs)
        t = doc["targets"][0]
        self.assertEqual(t["action"], "check_cost_model")
        self.assertTrue(t["flags"])

    def test_replay_far_faster_than_the_profile_is_flagged(self):
        recs = [_rec("suspicious", 1.0, traced=100.0, comparable=True,
                     nbytes=1_000)]
        doc = rank.rank(recs)
        self.assertTrue(any("faster than the profiled" in f
                            for f in doc["targets"][0]["flags"]))

    def test_only_the_chosen_operating_point_is_ranked(self):
        recs = [_rec("op", 10.0, batch_size=32), _rec("op", 10.0, batch_size=32),
                _rec("op", 999.0, batch_size=1)]
        doc = rank.rank(recs)
        self.assertEqual(doc["operating_points"]["decode"]["batch_size"], 32)
        self.assertEqual(doc["targets"][0]["e2e_us"], 20.0)

    def test_a_sweep_invariant_case_is_ranked_at_every_point_it_covers(self):
        # measured once at bs=1 but valid for bs=32 as well
        invariant = _rec("moe_grouped_gemm", 100.0, nbytes=1_000,
                         batch_size=1,
                         points=[["decode", 1, 2048, 1], ["decode", 1, 2048, 32]])
        varying = _rec("per_token_op", 10.0, nbytes=1_000, batch_size=32,
                       points=[["decode", 1, 2048, 32]])
        doc = rank.rank([invariant, varying, dict(varying)])
        self.assertEqual(doc["operating_points"]["decode"]["batch_size"], 32)
        self.assertIn("moe_grouped_gemm", [t["op"] for t in doc["targets"]])

    def test_ranking_without_any_measurement_is_an_error_not_an_empty_table(self):
        with self.assertRaises(ValueError):
            rank.rank([_rec("op", 1.0, status="failed")])


class TestTimingPlan(unittest.TestCase):
    def test_a_fast_kernel_is_repeated_inside_the_window(self):
        reps, windows = timing.plan_window(5e-6)      # 5 us
        self.assertGreater(reps, 100)
        self.assertGreaterEqual(windows, timing.MIN_WINDOWS)

    def test_a_slow_kernel_is_not_repeated_into_a_huge_window(self):
        reps, _ = timing.plan_window(0.05)            # 50 ms
        self.assertEqual(reps, timing.MIN_REPS)

    def test_restorer_is_only_built_for_mutated_operands(self):
        self.assertIsNone(timing.make_restorer([]))


class TestEstimate(unittest.TestCase):
    def test_timeout_covers_startup_measurement_and_allocation(self):
        t = estimate.op_timeout(10, 0.5, startup_s=60, case_overhead_s=3,
                                safety=3, alloc_bytes=20e9)
        # 60 + 10*3.5 + 10 s of allocation, times 3
        self.assertGreater(t, 300)
        self.assertLessEqual(t, estimate.MAX_TIMEOUT_S)

    def test_timeout_never_drops_below_the_floor(self):
        self.assertGreaterEqual(estimate.op_timeout(1, 0.01, startup_s=1,
                                                    case_overhead_s=0,
                                                    safety=1),
                                estimate.MIN_TIMEOUT_S)

    def test_calibration_uses_previous_runs(self):
        prev = [{"ops": [{"cases": 1, "seconds": 20.0},
                         {"cases": 11, "seconds": 130.0}]}]
        startup, per_case = estimate.calibrate(prev)
        self.assertAlmostEqual(startup, 20.0)
        self.assertAlmostEqual(per_case, 10.0)

    def test_per_op_budgets_are_carried_through_the_timeout_plan(self):
        """A per-op budget must size that op's timeout, not a global constant."""
        plan = estimate.plan({"slow": 200, "fast": 200},
                             {"slow": 2.0, "fast": 0.1},
                             safety=1.0)
        self.assertGreater(plan["slow"], plan["fast"])


class TestAdaptiveBudget(unittest.TestCase):
    """The measurement budget is derived from the profiled shape.

    There is no user-facing "budget / case" knob: how long a case needs is a
    property of the kernel being replayed, and the profile already knows it.
    """

    @staticmethod
    def _case(**kw):
        return spec.BenchCase(op=kw.pop("op", "aten::mm"), **kw)

    def test_a_slow_shape_gets_a_bigger_budget_than_a_fast_one(self):
        fast = self._case(traced_device_time_us=5.0, traced_comparable=True)
        slow = self._case(traced_device_time_us=50_000.0,
                          traced_comparable=True)
        self.assertLess(estimate.case_budget(fast, CACHE_PEAKS),
                        estimate.case_budget(slow, CACHE_PEAKS))

    def test_the_budget_stays_inside_its_bounds(self):
        """A microsecond kernel still buys full windows; a huge one is capped.

        The floor is the device-event window target (a shorter budget would buy
        fewer windows than the timer needs), and the ceiling stops one
        pathological shape from owning the run.
        """
        tiny = self._case(traced_device_time_us=0.001, traced_comparable=True)
        huge = self._case(traced_device_time_us=1e9, traced_comparable=True)
        self.assertGreaterEqual(estimate.case_budget(tiny, CACHE_PEAKS),
                                estimate.MIN_BUDGET_S)
        self.assertAlmostEqual(
            estimate.case_budget(tiny, CACHE_PEAKS),
            estimate.TARGET_WINDOWS * timing.TARGET_WINDOW_S)
        self.assertAlmostEqual(estimate.case_budget(huge, CACHE_PEAKS),
                               estimate.MAX_BUDGET_S)

    def test_a_case_without_a_traced_time_is_predicted_from_its_work(self):
        """No trace-comparable shape: fall back to the analytic cost."""
        light = self._case(nbytes=1e5)
        heavy = self._case(flops=1e13, nbytes=1e9)
        self.assertLess(estimate.case_budget(light, CACHE_PEAKS),
                        estimate.case_budget(heavy, CACHE_PEAKS))

    def test_an_op_is_budgeted_by_its_most_expensive_case(self):
        cheap = self._case(traced_device_time_us=5.0, traced_comparable=True)
        dear = self._case(traced_device_time_us=20_000.0,
                          traced_comparable=True)
        budgets = estimate.op_budgets({"aten::mm": [cheap, dear]}, CACHE_PEAKS)
        self.assertAlmostEqual(budgets["aten::mm"],
                               estimate.case_budget(dear, CACHE_PEAKS))

    def test_roofline_bound_is_the_fastest_possible_time(self):
        us, bound = estimate.roofline_bound_us(0.0, 456_000, PEAKS)
        self.assertEqual(bound, "memory")
        self.assertAlmostEqual(us, 1.0, places=2)

    def test_roofline_bound_uses_the_cache_roof_when_it_applies(self):
        us, bound = estimate.roofline_bound_us(0.0, 1_200_000, CACHE_PEAKS)
        self.assertEqual(bound, "memory")
        self.assertAlmostEqual(us, 1.0, places=2)


class TestReports(unittest.TestCase):
    def test_enrich_adds_utilization_and_the_fidelity_ratio(self):
        rich = reports.enrich([_rec("op", 2.0, nbytes=456_000, traced=4.0,
                                    comparable=True)], PEAKS)[0]
        self.assertAlmostEqual(rich["util"], 0.5, places=2)
        self.assertEqual(rich["memory_level"], "dram")
        self.assertAlmostEqual(rich["replay_vs_traced"], 0.5, places=2)

    def test_coverage_lists_what_was_not_measured_and_why(self):
        recs = [_rec("ok_op", 1.0),
                _rec("wrapper", 0.0, status="not_replayable",
                     detail="needs forward context", traced=60.0)]
        cov = reports.coverage(reports.enrich(recs, PEAKS))
        self.assertEqual(len(cov), 1)
        self.assertEqual(cov[0]["Op"], "wrapper")
        self.assertEqual(cov[0]["Reason"], "needs forward context")


class TestHardwareUnitRoofs(unittest.TestCase):
    """The roof is a *hardware unit*, and a vector op cannot reach XMX."""

    UNIT_PEAKS = dict(CACHE_PEAKS, vector_tflops=12.3, matrix_unit="XMX",
                      vector_unit="XVE", cache_name="L3-Cache")

    def test_matrix_family_ops_are_recognized(self):
        for op in ("aten::mm", "aten::bmm", "aten::addmm", "aten::linear",
                   "aten::_scaled_mm", "_moe_C::grouped_gemm",
                   "vllm::unified_attention_with_output"):
            self.assertTrue(estimate.uses_matrix_engine(op), op)
        for op in ("_C::rms_norm", "_C::silu_and_mul", "aten::embedding",
                   "_moe_C::moe_gather", "c10d::allreduce_"):
            self.assertFalse(estimate.uses_matrix_engine(op), op)

    def test_a_vector_op_is_scored_against_the_vector_peak(self):
        # A norm issues vector instructions; charging it to the 98.3 TFLOPS XMX
        # peak made every elementwise kernel look like it had ~99 % headroom.
        peak, unit = estimate.compute_peak(self.UNIT_PEAKS, "_C::rms_norm")
        self.assertEqual((peak, unit), (12.3, "XVE"))
        peak, unit = estimate.compute_peak(self.UNIT_PEAKS, "aten::linear")
        self.assertEqual((peak, unit), (98.3, "XMX"))
        # ... and the ridge point moves with it.
        self.assertLess(estimate.ridge_ai(self.UNIT_PEAKS, 456.0,
                                          "_C::rms_norm"),
                        estimate.ridge_ai(self.UNIT_PEAKS, 456.0,
                                          "aten::linear"))

    def test_the_roof_is_named_as_a_unit(self):
        d = estimate.roofline_detail(1.0, 0.0, 64 * 1024 ** 2, self.UNIT_PEAKS,
                                     "_C::rms_norm")
        self.assertEqual(d["unit"], "DRAM")
        d = estimate.roofline_detail(1.0, 0.0, 1024 ** 2, self.UNIT_PEAKS,
                                     "_C::rms_norm")
        self.assertEqual(d["unit"], "L3-Cache")
        # 300 MFLOP in 10 us on a vector op: AI 300 >> the XVE ridge.
        d = estimate.roofline_detail(10.0, 300e6, 1e6, self.UNIT_PEAKS,
                                     "_C::silu_and_mul")
        self.assertEqual((d["bound"], d["unit"]), ("compute", "XVE"))

    def test_a_cache_resident_op_is_also_scored_against_dram(self):
        # Cache-resident ops are usually already-optimal streaming kernels:
        # scored only against the (2.6x higher) cache roof they look like they
        # have headroom the model can never use. ``effective_util`` keeps the
        # honest cache number *and* the DRAM one.
        nbytes = 456_000  # 1 us of DRAM peak, and it fits the LLC
        d = estimate.roofline_detail(1.0, 0.0, nbytes, self.UNIT_PEAKS,
                                     "_C::rms_norm")
        self.assertEqual(d["memory_level"], "cache")
        self.assertAlmostEqual(d["util"], 456.0 / 1200.0, places=2)
        self.assertAlmostEqual(d["util_dram"], 1.0, places=2)
        self.assertAlmostEqual(d["effective_util"], 1.0, places=2)

    def test_an_op_at_the_dram_roof_is_not_a_kernel_target(self):
        # 1 us moving 456 kB: 38 % of the cache roof but 100 % of DRAM. The
        # cache headroom is unreachable, so this is not a kernel session.
        recs = [_rec("_C::rms_norm", 1.0, layers=36, nbytes=456_000)]
        doc = rank.rank(recs)
        t = doc["targets"][0]
        self.assertEqual(t["roofline"]["unit"], "L3-Cache")
        self.assertLess(t["roofline"]["util"], 0.5)
        self.assertGreaterEqual(t["roofline"]["effective_util"], 0.9)
        self.assertEqual(t["action"], "at_roofline")
        self.assertEqual(t["savings_us"]["total"], 0.0)
        self.assertTrue(any("DRAM roof" in f for f in t["flags"]))


class TestPhaseSeparation(unittest.TestCase):
    """Prefill and decode are different machines for the same kernel."""

    def test_targets_are_also_ranked_per_phase(self):
        recs = [
            _rec("_C::rms_norm", 10.0, layers=36, nbytes=1_000,
                 phase="prefill", seq_len=1024, ctx_len=0, batch_size=1),
            _rec("aten::linear", 500.0, layers=1, nbytes=1_000,
                 phase="prefill", seq_len=1024, ctx_len=0, batch_size=1),
            _rec("_C::rms_norm", 5.0, layers=36, nbytes=1_000, phase="decode"),
        ]
        doc = rank.rank(recs)
        self.assertIn("by_phase", doc)
        self.assertEqual(sorted(doc["by_phase"]), ["decode", "prefill"])
        # decode saw only one op; prefill saw both
        self.assertEqual([t["op"] for t in doc["by_phase"]["decode"]["targets"]],
                         ["_C::rms_norm"])
        self.assertEqual(len(doc["by_phase"]["prefill"]["targets"]), 2)
        # the per-phase e2e sums only that phase's weighted time
        self.assertAlmostEqual(doc["by_phase"]["decode"]["e2e_us_total"],
                               5.0 * 36, places=1)
        self.assertEqual(doc["by_phase"]["prefill"]["operating_point"],
                         {"seq_len": 1024, "ctx_len": 0, "batch_size": 1,
                          "profiled": False})
        # and the combined ranking is still there
        self.assertEqual(len(doc["targets"]), 2)

    def test_format_table_can_print_one_phase(self):
        recs = [_rec("_C::rms_norm", 5.0, layers=36, nbytes=1_000)]
        doc = rank.rank(recs)
        text = rank.format_table(doc, phase="decode")
        self.assertIn("=== decode ===", text)
        self.assertNotIn("=== prefill + decode ===", text)


class TestPerOpSheets(unittest.TestCase):
    """One sheet per op: a flat Cases sheet mixed shape vocabularies."""

    def test_sheet_names_are_legal_and_unique(self):
        used: set[str] = set()
        self.assertEqual(reports.sheet_name("vllm::xpu_topk_topp_sampler",
                                            used),
                         "vllm.xpu_topk_topp_sampler")
        long = reports.sheet_name(
            "vllm::unified_attention_with_output_and_more", used)
        self.assertLessEqual(len(long), 31)
        again = reports.sheet_name(
            "vllm::unified_attention_with_output_and_more", used)
        self.assertNotEqual(long, again)
        self.assertNotIn(":", again)

    def test_workbook_has_a_sheet_per_op(self):
        recs = [_rec("_C::rms_norm", 5.0, nbytes=1_000, phase="prefill"),
                _rec("_C::rms_norm", 3.0, nbytes=1_000),
                _rec("aten::linear", 50.0, nbytes=1_000)]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            try:
                reports.write_workbook(recs, path, CACHE_PEAKS,
                                       rank.rank(recs))
            except ImportError:  # pragma: no cover - pandas not installed
                self.skipTest("pandas/openpyxl not available")
            import openpyxl
            names = openpyxl.load_workbook(path).sheetnames
        self.assertIn("_C.rms_norm", names)
        self.assertIn("aten.linear", names)
        self.assertNotIn("Cases", names)
        self.assertIn("Targets prefill", names)
        self.assertIn("Targets decode", names)


class TestShapeMatrixInReport(unittest.TestCase):
    """The Shape Matrix is the run's *input*, so it ships in the run's report.

    Keeping it as a separate download made the reader correlate two files by
    hand: the measured latencies are only interpretable against the shapes they
    were measured at, and those shapes are what the cases were built from.
    """

    def _workbook(self, matrix):
        recs = [_rec("aten::linear", 50.0, nbytes=1_000, phase="prefill")]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "report.xlsx")
            try:
                reports.write_workbook(recs, path, CACHE_PEAKS, None, matrix)
            except ImportError:  # pragma: no cover - pandas not installed
                self.skipTest("pandas/openpyxl not available")
            import openpyxl
            wb = openpyxl.load_workbook(path)
            return wb.sheetnames, {n: [[c.value for c in r]
                                       for r in wb[n].iter_rows()]
                                   for n in wb.sheetnames}

    def test_matrix_and_info_sheets_are_written(self):
        rows = [{"Phase": "prefill", "Op Name": "aten::linear", "Seq Len": 2048,
                 "Ctx Len": 0, "Batch Size": 1, "TP": 1}]
        names, cells = self._workbook(
            {"info": [["Model", "Qwen/Qwen3-4B"]], "rows": rows})
        self.assertIn("Shape Matrix", names)
        self.assertIn("Info", names)
        # Info first (what the run is), Shape Matrix last (the longest sheet).
        self.assertEqual(names[0], "Info")
        self.assertEqual(names[-1], "Shape Matrix")
        matrix = cells["Shape Matrix"]
        self.assertEqual(matrix[0][0], "Phase")
        self.assertEqual(matrix[1][0], "prefill")
        # the row's values land under their own headers, not by position
        hdr = matrix[0]
        self.assertEqual(matrix[1][hdr.index("Seq Len")], 2048)
        self.assertEqual(cells["Info"][1], ["Model", "Qwen/Qwen3-4B"])

    def test_report_without_a_matrix_still_writes(self):
        names, _ = self._workbook(None)
        self.assertNotIn("Shape Matrix", names)
        self.assertNotIn("Info", names)
        self.assertIn("Summary", names)


class TestHistory(unittest.TestCase):
    def test_only_shapes_present_in_both_runs_are_compared(self):
        with tempfile.TemporaryDirectory() as d:
            conn = history.connect(history.db_path(d))
            history.ingest(conn, {"run_id": "base"},
                           [_rec("op", 10.0), _rec("gone", 5.0)])
            history.ingest(conn, {"run_id": "new"},
                           [_rec("op", 15.0), _rec("fresh", 5.0)])
            diffs = history.compare(conn, "base", "new")
            self.assertEqual([d_["op"] for d_ in diffs], ["op"])
            self.assertEqual(diffs[0]["kind"], "regression")
            self.assertAlmostEqual(diffs[0]["delta_pct"], 50.0)

    def test_failed_cases_are_not_ingested(self):
        with tempfile.TemporaryDirectory() as d:
            conn = history.connect(history.db_path(d))
            n = history.ingest(conn, {"run_id": "r"},
                               [_rec("op", 0.0, status="failed")])
            self.assertEqual(n, 0)


if __name__ == "__main__":
    unittest.main()


class TestAnalyticCostModel(unittest.TestCase):
    """The roofline is only as honest as the FLOPs/bytes it is fed."""

    def test_attention_has_analytic_work(self):
        # Attention is normally the heaviest op in the profile. With no cost
        # model it had zero FLOPs and zero bytes, so its bound came out
        # "unknown" and it was ranked as if it had 100 % headroom.
        from breakdown.cost import estimate_flops

        # decode: 32 queries, 32x2048 total KV rows, 32 heads x 128
        flops = estimate_flops("vllm::unified_attention_with_output",
                               [[32, 32, 128], [65536, 4, 128],
                                [65536, 4, 128], [32, 32, 128], [0]],
                               n_seqs=32)
        # each of the 32 queries attends its own 2048-token context
        self.assertEqual(flops, 2 * 2 * 32 * 2048 * 32 * 128)

    def test_an_empty_operand_does_not_zero_the_memory_estimate(self):
        # vLLM dispatches attention with an empty kv_cache_dummy_dep tensor
        # purely to order it against the KV write.
        from breakdown.shape_derive import _profile_op_memory

        with_dummy = _profile_op_memory(
            "vllm::unified_attention_with_output",
            [[32, 32, 128], [65536, 4, 128], [0]],
            ["bfloat16", "bfloat16", "bfloat16"], 2)
        self.assertGreater(with_dummy, 65536 * 4 * 128 * 2)

    def test_a_table_lookup_is_charged_for_the_rows_it_reads(self):
        # Charging an embedding for the whole vocabulary matrix produced a
        # "37000 % of peak" utilization that said nothing about the kernel.
        from breakdown.shape_derive import _profile_op_memory

        few = _profile_op_memory("aten::embedding", [[151936, 2560], [32]],
                                 ["bfloat16", "long int"], 2)
        self.assertLess(few, 151936 * 2560 * 2 / 100)
        self.assertGreaterEqual(few, 32 * 2560 * 2)

        rope = _profile_op_memory(
            "_C::rotary_embedding",
            [[32], [32, 4096], [32, 1024], [262144, 128]],
            ["long int", "bfloat16", "bfloat16", "bfloat16"], 2)
        self.assertLess(rope, 262144 * 128 * 2 / 100)

    def test_a_full_table_read_is_left_alone(self):
        from breakdown.shape_derive import _profile_op_memory

        # more indices than rows: the whole table really is streamed
        self.assertEqual(
            _profile_op_memory("aten::embedding", [[8, 4], [64]],
                               ["bfloat16", "long int"], 2),
            8 * 4 * 2 + 64 * 8 + 8 * 4 * 2)


class TestProfiledOperatingPoint(unittest.TestCase):
    """The ranking defaults to the point the *profile* ran, not the busiest.

    A sweep measures dozens of what-if points; only the one whose shapes equal
    the trace's is answerable against the profile (and against
    ``traced_device_time_us``). Ranking at the busiest swept point instead
    silently reported a shape the model never ran.
    """

    def test_profiled_point_wins_over_the_busiest_swept_point(self):
        recs = [
            # three cases at a swept point, one at the profiled point
            _rec("_C::rms_norm", 5.0, batch_size=1),
            _rec("aten::linear", 6.0, batch_size=1),
            _rec("aten::mm", 7.0, batch_size=1),
            _rec("_C::rms_norm", 9.0, batch_size=32, comparable=True,
                 traced=9.0),
        ]
        self.assertEqual(rank.profiled_point(recs, "decode"), (1, 2048, 32))
        self.assertEqual(rank.pick_point(recs, "decode"), (1, 2048, 32))
        doc = rank.rank(recs)
        self.assertEqual(doc["operating_points"]["decode"],
                         {"seq_len": 1, "ctx_len": 2048, "batch_size": 32,
                          "profiled": True})
        # only the profiled case is ranked
        self.assertEqual([t["op"] for t in doc["targets"]], ["_C::rms_norm"])

    def test_busiest_point_is_the_fallback_and_is_marked_unprofiled(self):
        recs = [_rec("_C::rms_norm", 5.0, batch_size=1),
                _rec("aten::mm", 6.0, batch_size=1)]
        self.assertIsNone(rank.profiled_point(recs, "decode"))
        doc = rank.rank(recs)
        self.assertFalse(doc["operating_points"]["decode"]["profiled"])

    def test_explicit_point_still_wins(self):
        recs = [_rec("_C::rms_norm", 5.0, batch_size=1),
                _rec("_C::rms_norm", 9.0, batch_size=32, comparable=True)]
        rc = rank.RankConfig(points={"decode": (1, 2048, 1)})
        self.assertEqual(rank.pick_point(recs, "decode", rc.points["decode"]),
                         (1, 2048, 1))

    def test_profiled_shape_is_listed_first_for_the_target(self):
        recs = [_rec("_C::rms_norm", 5.0, layers=36, batch_size=32,
                     case_id="swept", shape="[1, 6144]"),
                _rec("_C::rms_norm", 1.0, layers=1, batch_size=32,
                     comparable=True, traced=1.0, case_id="profiled",
                     shape="[32, 6144]")]
        doc = rank.rank(recs)
        shapes = doc["targets"][0]["top_shapes"]
        self.assertTrue(shapes[0]["profiled"])
        self.assertEqual(shapes[0]["shape"], "[32, 6144]")
        self.assertFalse(shapes[1]["profiled"])
